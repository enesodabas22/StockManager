import sys
import os
import sqlite3
import hashlib
import csv
import json
import math
from datetime import datetime, date

import certifi

os.environ['SSL_CERT_FILE'] = certifi.where()

"""
Proje: StockFlow Stok Takip Sistemi
Sürüm: 2.1.0

Açıklama:
    Bu modül, StockFlow uygulamasının ana giriş noktasıdır. PyQt6 kütüphanesi kullanılarak
    geliştirilen bu masaüstü uygulaması, küçük ve orta ölçekli işletmeler için
    kapsamlı stok takibi, satış yönetimi ve raporlama özellikleri sunar.

    Temel Özellikler:
    - SQLite tabanlı yerel veritabanı yönetimi
    - Firebase Realtime Database ile bulut yedekleme ve senkronizasyon
    - Dinamik grafikler ve veri görselleştirme (Matplotlib alternatifi custom widgetlar)
    - Kullanıcı yetkilendirme ve işlem loglama sistemi
    - Excel/CSV formatında raporlama
"""

# --- ÜÇÜNCÜ PARTİ KÜTÜPHANELER ---
try:
    from plyer import notification

    PLYER_MUMKUN = True
except ImportError:
    PLYER_MUMKUN = False
    print("UYARI: 'plyer' kütüphanesi bulunamadı. Masaüstü bildirimleri çalışmayacak.")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_MUMKUN = True
except ImportError:
    EXCEL_MUMKUN = False
    print("UYARI: 'openpyxl' bulunamadı. Excel çıktısı alınamayacak.")

try:
    import firebase_admin
    from firebase_admin import credentials, db

    FIREBASE_MUMKUN = True
except ImportError:
    FIREBASE_MUMKUN = False
    print("UYARI: 'firebase-admin' bulunamadı. Bulut yedekleme çalışmayacak.")
    print("Kurmak için: pip install firebase-admin")

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QMessageBox, QTableWidget, QTableWidgetItem, QMenu, QDialog,
    QFormLayout, QDialogButtonBox, QAbstractItemView, QFileDialog,
    QMainWindow, QMenuBar, QCheckBox, QHeaderView, QFrame, QStackedWidget,
    QListWidget, QListWidgetItem, QStatusBar, QInputDialog, QComboBox,
    QProgressBar, QDateEdit, QGraphicsDropShadowEffect, QScrollArea, QTabWidget, QTabBar,
    QGraphicsOpacityEffect
)
from PyQt6.QtGui import QAction, QFont, QColor, QCursor, QPixmap, QDoubleValidator, QPainter, QPen, QBrush, \
    QLinearGradient, QRegularExpressionValidator, QPainterPath
from PyQt6.QtCore import Qt, pyqtSignal, QPoint, QDate, QRectF, QSize, QLocale, QRegularExpression, \
    QPropertyAnimation, QEasingCurve, pyqtProperty, QParallelAnimationGroup

STANDART_BIRIMLER = [
    "Adet", "Kg", "Paket", "Kutu", "Koli",
    "Çuval", "Palet", "Ton", "Kasa"]


# --- YAPILANDIRMA ---
def dosya_yolunu_bul(relative_path):
    """ PyInstaller ile paketlenmiş exe ve normal geliştirme ortamı için doğru yolu döner """
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller geçici klasörü (Paketlenmiş uygulama buraya bakar)
        return os.path.join(sys._MEIPASS, relative_path)
    # Normal çalışma ortamı (PyCharm vb.)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), relative_path)


# Artık dosya yolunu bu fonksiyonla alıyoruz
FIREBASE_KEY_PATH = dosya_yolunu_bul("firebase_key.json")
FIREBASE_DB_URL = "https://stockfloww-3cf71-default-rtdb.europe-west1.firebasedatabase.app/"
DB_NAME = "stok_veritabani.db"


# =============================================================================
# 1. YARDIMCI SINIFLAR VE VERİTABANI YÖNETİMİ
# =============================================================================

class NumericTableWidgetItem(QTableWidgetItem):
    """
    Kullanıcı arayüzünde sayısal verilerin (Fiyat, Miktar vb.) doğru sıralanması
    için kullanılan özelleştirilmiş tablo hücresi sınıfıdır.

    Normal QTableWidgetItem metin tabanlı sıralama yapar (Örn: "100" < "20"),
    bu sınıf ise sayısal büyüklüğe göre sıralamayı garanti eder.
    """

    def __init__(self, display_text, sort_key):
        """
        Yapıcı Metot (Constructor).

        Args:
            display_text (str): Ekranda görünecek metin (örn: "120.50₺").
            sort_key (float/int): Sıralamada baz alınacak gerçek sayısal değer (örn: 120.50).
        """
        super().__init__(display_text)
        self.sort_key = sort_key

    def __lt__(self, other):
        """
        Küçüktür (<) operatörünün aşırı yüklenmesi (Operator Overloading).
        Python'un sıralama algoritmaları (sort) tarafından kullanılır.
        """
        if isinstance(other, NumericTableWidgetItem):
            return self.sort_key < other.sort_key
        return super().__lt__(other)


class VeritabaniYoneticisi:
    """
    Veritabanı Yönetim Sınıfı (Database Manager)

    Uygulamanın SQLite veritabanı ile olan tüm etkileşimlerini (CRUD işlemleri) yönetir.
    Singleton tasarım desenine benzer şekilde, uygulama genelinde veritabanı erişimi
    bu sınıf üzerinden sağlanır.

    Görevleri:
    - Veritabanı bağlantısını kurmak ve sürdürmek.
    - Tablo yapısını oluşturmak ve versiyon geçişlerini (migrasyon) yönetmek.
    - Ürün, stok, kullanıcı ve raporlama sorgularını işlemek.
    """

    def __init__(self, db_adi=DB_NAME):
        """
        Veritabanı bağlantısını başlatır ve gerekli tabloları kontrol eder.

        Args:
            db_adi (str): Veritabanı dosyasının adı. Varsayılan: stok_veritabani.db
        """
        self.baglanti = sqlite3.connect(db_adi, detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES)
        # Foreign Key desteğini aktifleşitir
        self.baglanti.execute("PRAGMA foreign_keys = ON")
        self.cursor = self.baglanti.cursor()
        self.tablolari_olustur()
        self.veritabani_migrasyonu_kontrol_et()

    def _sutun_tipi_getir(self, tablo_adi, sutun_adi):
        try:
            self.cursor.execute(f"PRAGMA table_info({tablo_adi})")
            for row in self.cursor.fetchall():
                if row[1] == sutun_adi:
                    return row[2]
            return None
        except sqlite3.OperationalError:
            return None

    def veritabani_migrasyonu_kontrol_et(self):
        try:
            self.cursor.execute("PRAGMA table_info(urunler)")
            mevcut_sutunlar = [row[1] for row in self.cursor.fetchall()]

            # 'aktif' sütunu
            if 'aktif' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN aktif INTEGER DEFAULT 1")


            if 'urun_kodu' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN urun_kodu TEXT")
                try:
                    self.cursor.execute("UPDATE urunler SET urun_kodu = 'KOD-' || id WHERE urun_kodu IS NULL")
                    self.cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_urun_kodu ON urunler(urun_kodu)")
                except sqlite3.IntegrityError:
                    pass

            if 'birim' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN birim TEXT NOT NULL DEFAULT 'adet'")
            if 'baslangic_miktari' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN baslangic_miktari REAL")
                self.cursor.execute("UPDATE urunler SET baslangic_miktari = miktar WHERE baslangic_miktari IS NULL")
            if 'son_kullanma_tarihi' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN son_kullanma_tarihi TEXT")

            self.cursor.execute("PRAGMA table_info(stok_hareketleri)")
            hareket_sutunlari = [row[1] for row in self.cursor.fetchall()]
            if 'satis_fiyati' not in hareket_sutunlari:
                self.cursor.execute("ALTER TABLE stok_hareketleri ADD COLUMN satis_fiyati REAL")

            # resim yolu sütunu
            if 'resim_yolu' not in mevcut_sutunlar:
                self.cursor.execute("ALTER TABLE urunler ADD COLUMN resim_yolu TEXT")


            self.baglanti.commit()
        except sqlite3.OperationalError:
            pass

    def tablolari_olustur(self):
        """
        Veritabanı tablolarını yoksa oluşturur (IF NOT EXISTS).

        Tablolar:
        1. urunler: Stoktaki ürünlerin detayları (master tablo).
        2. kullanicilar: Sisteme giriş yapacak yetkili hesaplar.
        3. stok_hareketleri: Ürün giriş-çıkış logları (transaction log).
        """
        # Ürünler tablosu
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS urunler (
                id INTEGER PRIMARY KEY,
                urun_kodu TEXT UNIQUE NOT NULL,      -- Benzersiz Barkod/Kod
                ad TEXT NOT NULL,
                kategori TEXT,
                fiyat REAL NOT NULL DEFAULT 0.0,
                miktar REAL NOT NULL,
                birim TEXT NOT NULL DEFAULT 'adet',
                min_stok REAL NOT NULL DEFAULT 10,  -- Kritik stok uyarısı için eşik
                baslangic_miktari REAL,              -- Doluluk barı hesabı için
                son_kullanma_tarihi TEXT,
                resim_yolu TEXT,
                aktif INTEGER DEFAULT 1              -- Soft delete için (1: Aktif, 0: Silinmiş)
            )""")

        # Kullanıcılar tablosu
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS kullanicilar (
                id INTEGER PRIMARY KEY,
                kullanici_adi TEXT NOT NULL UNIQUE,
                sifre_hash TEXT NOT NULL             -- SHA-256 ile hashlenmiş şifre
            )""")

        # Stok Hareketleri (Log) tablosu
        self.cursor.execute("""
            CREATE TABLE IF NOT EXISTS stok_hareketleri (
                id INTEGER PRIMARY KEY,
                urun_id INTEGER,
                kullanici_adi TEXT,
                islem_tipi TEXT NOT NULL,            -- Örn: 'STOK EKLEME', 'STOK ÇIKIŞI' 
                miktar_degisimi REAL NOT NULL,
                yeni_miktar REAL,
                tarih TIMESTAMP,
                notlar TEXT,
                satis_fiyati REAL,
                FOREIGN KEY (urun_id) REFERENCES urunler(id) ON DELETE SET NULL
            )""")
        self.baglanti.commit()

    def _stok_hareketi_kaydet(self, urun_id, kullanici_adi, islem_tipi, miktar_degisimi, yeni_miktar, notlar="",
                              satis_fiyati=None):
        """
        Dahili Metot: Bir stok işlemini log tablosuna kaydeder.
        Bu metot, veri bütünlüğü için her ekleme/silme işleminden sonra otomatik çağrılır.
        """
        try:
            guncel_tarih = datetime.now()
            self.cursor.execute("""
                INSERT INTO stok_hareketleri (urun_id, kullanici_adi, islem_tipi, miktar_degisimi, yeni_miktar, notlar, tarih, satis_fiyati)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (urun_id, kullanici_adi, islem_tipi, miktar_degisimi, yeni_miktar, notlar, guncel_tarih, satis_fiyati))
        except Exception as e:
            print(f"Stok hareketi kaydedilemedi: {e}")

    def urunleri_getir(self):
        """Aktif (silinmemiş) tüm ürünlerin listesini ürün adına göre sıralı getirir."""
        # Sadece aktif=1 olanları getiriyoruz (Soft delete filtresi)
        self.cursor.execute(
            "SELECT id, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, baslangic_miktari, son_kullanma_tarihi FROM urunler WHERE aktif = 1 ORDER BY ad ASC")
        return self.cursor.fetchall()

    def dusuk_stok_urunleri_getir(self):
        """Stok miktarı, minimum stok seviyesinin (eşik) altına düşen ürünleri raporlar."""
        self.cursor.execute(
            "SELECT id, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok FROM urunler WHERE miktar <= min_stok AND aktif = 1 ORDER BY ad ASC")
        return self.cursor.fetchall()

    def urun_detay_getir(self, urun_id):
        """Belirli bir ürünün ID'sine göre tüm detaylarını çeker."""
        self.cursor.execute(
            "SELECT id, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, son_kullanma_tarihi FROM urunler WHERE id = ?",
            (urun_id,))
        return self.cursor.fetchone()

    def urun_ekle(self, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, skt_tarihi, kullanici_adi):
        """
        Yeni bir stok kartı oluşturur veya mevcut ürün varsa stoğunu artırır.

        Args:
            urun_kodu (str): Ürüne ait benzersiz barkod veya kod.
            ad (str): Ürün adı.
            kategori (str): Ürün kategorisi.
            fiyat (float): Alış maliyeti.
            miktar (float): Eklenecek miktar.
            ...

        Returns:
            tuple: (Basari_Durumu (bool), Mesaj (str))
        """
        self.cursor.execute("SELECT id, miktar FROM urunler WHERE urun_kodu = ?", (urun_kodu,))
        mevcut = self.cursor.fetchone()

        if mevcut:
            # Ürün zaten varsa: Miktarını güncelle (UPDATE)
            urun_id, mevcut_miktar = mevcut
            yeni_miktar = mevcut_miktar + miktar
            self.cursor.execute(
                "UPDATE urunler SET miktar = ?, baslangic_miktari = baslangic_miktari + ?, aktif = 1 WHERE id = ?",
                (yeni_miktar, miktar, urun_id))
            self._stok_hareketi_kaydet(urun_id, kullanici_adi, "STOK EKLEME", miktar, yeni_miktar, "Formdan eklendi")
            self.baglanti.commit()
            return True, f"Mevcut '{urun_kodu}' kodlu ürünün stoğu güncellendi. (SKT Değişmedi)"
        else:
            # Ürün yoksa: Yeni kayıt oluştur (INSERT)
            try:
                self.cursor.execute("""
                    INSERT INTO urunler (urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, baslangic_miktari, son_kullanma_tarihi, aktif)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
                """, (urun_kodu, ad.upper(), kategori.upper(), fiyat, miktar, birim, min_stok, miktar, skt_tarihi))

                urun_id = self.cursor.lastrowid
                self._stok_hareketi_kaydet(urun_id, kullanici_adi, "STOK GİRİŞİ", miktar, miktar, "Yeni Ürün Ekleme")
                self.baglanti.commit()
                return True, "Ürün başarıyla eklendi."
            except sqlite3.IntegrityError:
                self.baglanti.rollback()
                return False, f"'{urun_kodu}' ürün kodu başkası tarafından kullanılıyor."
            except Exception as e:
                self.baglanti.rollback()
                return False, f"Bir hata oluştu: {e}"

    def urun_detay_guncelle(self, urun_id, yeni_ad, kategori, fiyat, birim, min_stok, yeni_skt, kullanici_adi):
        """Mevcut bir ürünün (stok miktarı hariç) diğer bilgilerini günceller."""
        try:
            self.cursor.execute("""
                UPDATE urunler 
                SET ad = ?, kategori = ?, fiyat = ?, birim = ?, min_stok = ?, son_kullanma_tarihi = ?
                WHERE id = ?
            """, (yeni_ad.upper(), kategori.upper(), fiyat, birim, min_stok, yeni_skt, urun_id))

            notlar = f"Detaylar güncellendi (Birim: {birim}, MinStok: {min_stok}, SKT: {yeni_skt})"
            self.cursor.execute("SELECT miktar FROM urunler WHERE id = ?", (urun_id,))
            result = self.cursor.fetchone()
            mevcut_miktar = result[0] if result else 0

            self._stok_hareketi_kaydet(urun_id, kullanici_adi, "DETAY GÜNCELLEME", 0, mevcut_miktar, notlar)
            self.baglanti.commit()
            return True, "Güncellendi"
        except Exception as e:
            self.baglanti.rollback()
            return False, f"Hata: {e}"

    def urun_miktar_guncelle(self, urun_id, miktar_farki, kullanici_adi, satis_fiyati=None):
        """
        Stok artırma veya azaltma işlemlerini yönetir.

        Args:
            miktar_farki: (+5 ekleme, -3 çıkarma gibi)
        """
        self.cursor.execute("SELECT miktar, baslangic_miktari FROM urunler WHERE id = ?", (urun_id,))
        sonuc = self.cursor.fetchone()
        if not sonuc: return False, "Ürün bulunamadı."

        mevcut_miktar, mevcut_baslangic = sonuc
        yeni_miktar = mevcut_miktar + miktar_farki

        if yeni_miktar < 0: return False, "Stok eksiye düşemez."

        # --- Soft Delete (Pasife Çekme) Kontrolü ---
        if yeni_miktar == 0:
            islem_tipi = "STOK ÇIKIŞI" if miktar_farki < 0 else "STOK GÜNCELLEME"
            # Notlara "OTOMATİK SİLİNDİ" yazıyoruz ki geri alırken anlayabilelim
            self._stok_hareketi_kaydet(urun_id, kullanici_adi, islem_tipi, miktar_farki, 0,
                                       "Stok bitti, ürün arşivlendi (Pasif).", satis_fiyati)

            # Stok 0 olunca ürünü pasife çekiyoruz (aktif=0)
            self.cursor.execute("UPDATE urunler SET miktar = 0, aktif = 0 WHERE id = ?", (urun_id,))
            self.baglanti.commit()
            return True, "Ürün tükendiği için listeden kaldırıldı (Geçmişte görünür)."

        # --- Normal Güncelleme ---
        yeni_baslangic = max(mevcut_baslangic, yeni_miktar)  # Barın bozulmaması için
        self.cursor.execute("UPDATE urunler SET miktar = ?, baslangic_miktari = ?, aktif = 1 WHERE id = ?",
                            (yeni_miktar, yeni_baslangic, urun_id))

        islem_tipi = "STOK EKLEME" if miktar_farki > 0 else "STOK ÇIKIŞI"
        self._stok_hareketi_kaydet(urun_id, kullanici_adi, islem_tipi, miktar_farki, yeni_miktar, "Manuel işlem",
                                   satis_fiyati)
        self.baglanti.commit()
        return True, "Miktar güncellendi."

    def urun_sil(self, urun_id, kullanici_adi):
        """
        Ürünü veritabanından tamamen silmek yerine 'soft delete' uygular.
        Ürün arşivlenir ve listelerde görünmez, ancak geçmiş hareketlerde referansı korunur.
        """
        mevcut_miktar = self.mevcut_miktar_getir(urun_id)
        self._stok_hareketi_kaydet(urun_id, kullanici_adi, "ÜRÜN SİLME", -mevcut_miktar, 0, "Kullanıcı sildi (Pasif)")
        # Tamamen silmek yerine pasife çekiyoruz
        self.cursor.execute("UPDATE urunler SET aktif = 0, miktar = 0 WHERE id = ?", (urun_id,))
        self.baglanti.commit()

    def mevcut_miktar_getir(self, urun_id):
        """Belirtilen ürünün anlık stok miktarını sorgular."""
        self.cursor.execute("SELECT miktar FROM urunler WHERE id = ?", (urun_id,))
        sonuc = self.cursor.fetchone()
        return sonuc[0] if sonuc else 0

    def tum_kullanicilari_getir(self):
        """Yedekleme için tüm kullanıcı bilgilerini getirir."""
        self.cursor.execute("SELECT id, kullanici_adi, sifre_hash FROM kullanicilar")
        return self.cursor.fetchall()

    def kullanici_yukle_raw(self, u_id, k_adi, s_hash):
        """Yedekten dönerken kullanıcıyı olduğu gibi (şifreyi tekrar hashlemeden) kaydeder."""
        try:
            self.cursor.execute("""
                INSERT INTO kullanicilar (id, kullanici_adi, sifre_hash) 
                VALUES (?, ?, ?)
            """, (u_id, k_adi, s_hash))
            self.baglanti.commit()
        except Exception as e:
            print(f"Kullanıcı yükleme hatası ({k_adi}): {e}")

    def urun_hucre_guncelle(self, urun_id, sutun_adi, yeni_deger, kullanici_adi):
        """
        Tablo üzerinden yapılan hızlı düzenlemeleri (Inline Editing) veritabanına işler.
        Sadece 'kategori' ve 'fiyat' alanlarının bu şekilde güncellenmesine izin verilir.
        """
        izin_verilen_sutunlar = ['kategori', 'fiyat']
        if sutun_adi not in izin_verilen_sutunlar: return False, "Geçersiz güncelleme alanı."
        try:
            self.cursor.execute(f"UPDATE urunler SET {sutun_adi} = ? WHERE id = ?", (yeni_deger, urun_id))
            self.cursor.execute("SELECT miktar FROM urunler WHERE id = ?", (urun_id,))
            mevcut_miktar = self.cursor.fetchone()[0]
            notlar = f"Hızlı düzenleme: {sutun_adi} -> {yeni_deger}"
            self._stok_hareketi_kaydet(urun_id, kullanici_adi, "DETAY GÜNCELLEME", 0, mevcut_miktar, notlar)
            self.baglanti.commit()
            return True, "Güncellendi."
        except Exception as e:
            self.baglanti.rollback()
            return False, f"Veritabanı hatası: {e}"

    def stok_hareketlerini_getir(self, zaman_araligi='tumu'):
        """
        Geçmiş stok işlemlerini (Logları) belirli bir zaman aralığına göre filtreleyerek getirir.

        Args:
            zaman_araligi (str): 'haftalik', 'aylik' veya 'tumu'.
        """
        sorgu = """
            SELECT h.id, h.tarih, h.kullanici_adi, u.urun_kodu, u.ad, h.islem_tipi, h.miktar_degisimi, h.yeni_miktar, u.birim, h.notlar, h.satis_fiyati
            FROM stok_hareketleri h
            LEFT JOIN urunler u ON h.urun_id = u.id
        """
        if zaman_araligi == 'haftalik':
            sorgu += " WHERE h.tarih >= date('now', '-7 day')"
        elif zaman_araligi == 'aylik':
            sorgu += " WHERE h.tarih >= date('now', '-30 day')"
        sorgu += " ORDER BY h.tarih DESC"
        self.cursor.execute(sorgu)
        return self.cursor.fetchall()

    def islem_geri_al(self, hareket_id, aktif_kullanici):
        """
        Yapılan hatalı bir stok işlemini geri alır (Undo).
        Örneğin yanlışlıkla 5 adet eklendiyse, bu işlem geri alınarak 5 adet düşülür.
        Eğer işlem sonucunda stok eksiye düşecekse geri almaya izin verilmez.
        """
        self.cursor.execute("SELECT urun_id, miktar_degisimi, islem_tipi FROM stok_hareketleri WHERE id = ?",
                            (hareket_id,))
        hareket = self.cursor.fetchone()
        if not hareket: return False, "İşlem bulunamadı."

        urun_id, miktar_degisimi, islem_tipi = hareket

        # Ürünü bulurken aktif/pasif fark etmeksizin buluyoruz
        self.cursor.execute("SELECT miktar FROM urunler WHERE id = ?", (urun_id,))
        urun = self.cursor.fetchone()

        if not urun: return False, "Ürün veritabanından tamamen silinmiş (Geri alınamaz)."

        mevcut_stok = urun[0]
        ters_degisim = -1 * miktar_degisimi
        yeni_miktar = mevcut_stok + ters_degisim

        if yeni_miktar < 0: return False, "Geri alma işlemi stok miktarını eksiye düşüreceği için yapılamaz."

        try:
            # Geri alırken aktif = 1 yaparak ürünü tekrar listeye sokuyoruz
            self.cursor.execute("UPDATE urunler SET miktar = ?, aktif = 1 WHERE id = ?", (yeni_miktar, urun_id))

            notlar = f"İşlem ID: {hareket_id} geri alındı."
            self._stok_hareketi_kaydet(urun_id, aktif_kullanici, "İŞLEM GERİ ALMA", ters_degisim, yeni_miktar, notlar)
            self.baglanti.commit()
            return True, "İşlem geri alındı ve ürün tekrar listeye eklendi."
        except Exception as e:
            self.baglanti.rollback()
            return False, f"Hata: {e}"

    def en_cok_satanlari_getir(self, zaman_araligi='aylik'):
        """En çok stok çıkışı yapılan (satılan) ürünleri çoktan aza sıralar."""
        zaman_kriteri = ""
        if zaman_araligi == 'haftalik':
            zaman_kriteri = "AND h.tarih >= date('now', '-7 day')"
        elif zaman_araligi == 'aylik':
            zaman_kriteri = "AND h.tarih >= date('now', '-30 day')"
        elif zaman_araligi == 'yillik':
            zaman_kriteri = "AND h.tarih >= date('now', '-365 day')"
        sorgu = f"""
            SELECT u.urun_kodu, u.ad, u.kategori, COALESCE(ABS(SUM(h.miktar_degisimi)), 0) as toplam_satis, u.birim
            FROM urunler u
            LEFT JOIN stok_hareketleri h ON u.id = h.urun_id AND h.islem_tipi = 'STOK ÇIKIŞI' {zaman_kriteri}
            GROUP BY u.id
            ORDER BY toplam_satis DESC
        """
        self.cursor.execute(sorgu)
        return self.cursor.fetchall()

    def kar_zarar_raporu_getir(self, zaman_araligi='aylik'):
        """
        Satılan ürünlerin alış fiyatı ile satış fiyatı arasındaki farkı hesaplayarak
        toplam ciro ve kar/zarar durumunu raporlar.
        """
        zaman_kriteri = ""
        if zaman_araligi == 'haftalik':
            zaman_kriteri = "AND h.tarih >= date('now', '-7 day')"
        elif zaman_araligi == 'aylik':
            zaman_kriteri = "AND h.tarih >= date('now', '-30 day')"
        elif zaman_araligi == 'yillik':
            zaman_kriteri = "AND h.tarih >= date('now', '-365 day')"
        sorgu = f"""
            SELECT u.urun_kodu, u.ad, u.kategori, SUM(ABS(h.miktar_degisimi)) as satis_adedi, u.fiyat as guncel_alis_fiyati, AVG(h.satis_fiyati) as ort_satis_fiyati, SUM(ABS(h.miktar_degisimi) * h.satis_fiyati) as toplam_ciro, SUM(ABS(h.miktar_degisimi) * (h.satis_fiyati - u.fiyat)) as toplam_kar
            FROM stok_hareketleri h
            JOIN urunler u ON h.urun_id = u.id
            WHERE h.islem_tipi = 'STOK ÇIKIŞI' AND h.satis_fiyati IS NOT NULL AND h.satis_fiyati > 0 {zaman_kriteri}
            GROUP BY u.id
            ORDER BY toplam_kar DESC
        """
        self.cursor.execute(sorgu)
        return self.cursor.fetchall()

    def veritabanini_sifirla(self, kapsam='tumu', kullanici_adi=None):
        """
        Veritabanındaki verileri temizler (Fabrika Ayarları).

        Args:
            kapsam (str):
                - 'gecmis': Sadece stok hareketlerini siler.
                - 'tumu': Ürünleri ve hareketleri siler.

        Not: Kullanıcı hesapları güvenlik nedeniyle silinmez.
        """
        try:
            # 1. Ana verileri sil
            if kapsam == 'gecmis':
                self.cursor.execute("DELETE FROM stok_hareketleri")
                not_mesaji = "Sadece işlem geçmişi temizlendi."

            elif kapsam == 'tumu':
                self.cursor.execute("DELETE FROM stok_hareketleri")
                self.cursor.execute("DELETE FROM urunler")
                not_mesaji = "Tüm ürünler ve geçmiş veriler silindi (Fabrika ayarları)."

            # 2. Sayaçları (ID'leri) sıfırlamayı dene
            # Bu kısım 'no such table: sqlite_sequence' hatasını engeller.
            try:
                self.cursor.execute("DELETE FROM sqlite_sequence WHERE name='stok_hareketleri'")
                if kapsam == 'tumu':
                    self.cursor.execute("DELETE FROM sqlite_sequence WHERE name='urunler'")
            except sqlite3.OperationalError:
                # sqlite_sequence tablosu yoksa (daha önce hiç veri girilmemişse) sorun yok, devam et.
                pass

            self.baglanti.commit()
            return True, not_mesaji

        except Exception as e:
            self.baglanti.rollback()
            return False, f"Sıfırlama hatası: {e}"

    def genel_bakis_getir(self):
        """Dashboard ekranı için özet metrikleri (toplam değer, düşük stok vb.) hesaplar."""
        try:
            urun_cesidi = self.cursor.execute("SELECT COUNT(id) FROM urunler").fetchone()[0]
            toplam_stok_degeri = self.cursor.execute("SELECT SUM(fiyat * miktar) FROM urunler").fetchone()[0]
            dusuk_stok_sayisi = \
                self.cursor.execute("SELECT COUNT(id) FROM urunler WHERE miktar <= min_stok").fetchone()[0]
            return {"urun_cesidi": urun_cesidi or 0, "toplam_deger": toplam_stok_degeri or 0.0,
                    "dusuk_stok": dusuk_stok_sayisi or 0}
        except Exception:
            return {"urun_cesidi": 0, "toplam_deger": 0.0, "dusuk_stok": 0}

    def kullanici_sayisi_getir(self):
        """Sisteme kayıtlı yetkili kullanıcı sayısını döner."""
        self.cursor.execute("SELECT COUNT(*) FROM kullanicilar")
        return self.cursor.fetchone()[0]

    def sifre_hashle(self, s):
        """Şifreleri veritabanında açık halde saklamamak için SHA-256 ile şifreler."""
        return hashlib.sha256(s.encode()).hexdigest()

    def kullanici_ekle(self, k_adi, s):
        """Yeni bir yönetici/kullanıcı hesabı oluşturur."""
        if not k_adi or not s: return False, "Kullanıcı adı ve şifre boş bırakılamaz."
        try:
            self.cursor.execute("INSERT INTO kullanicilar (kullanici_adi, sifre_hash) VALUES (?, ?)",
                                (k_adi, self.sifre_hashle(s)))
            self.baglanti.commit()
            return True, "Kullanıcı oluşturuldu."
        except sqlite3.IntegrityError:
            self.baglanti.rollback()
            return False, "Bu kullanıcı adı zaten alınmış."

    def kullanici_dogrula(self, k_adi, s):
        """Giriş ekranında girilen bilgilerin doğruluğunu kontrol eder."""
        self.cursor.execute("SELECT * FROM kullanicilar WHERE kullanici_adi = ? AND sifre_hash = ?",
                            (k_adi, self.sifre_hashle(s)))
        return self.cursor.fetchone() is not None

    def kullanici_bilgilerini_guncelle(self, e_kadi, y_kadi, y_sifre):
        """Mevcut kullanıcının adını veya şifresini değiştirir."""
        try:
            self.cursor.execute("UPDATE kullanicilar SET kullanici_adi = ?, sifre_hash = ? WHERE kullanici_adi = ?",
                                (y_kadi, self.sifre_hashle(y_sifre), e_kadi))
            self.baglanti.commit()
            return True, "Bilgiler güncellendi."
        except sqlite3.IntegrityError:
            self.baglanti.rollback()
            return False, "Yeni kullanıcı adı başkası tarafından kullanılıyor."


# =============================================================================
# 3. DIALOG WINDOWS
# =============================================================================

class FirebaseYedekleyici(QDialog):
    """Firebase Realtime Database ile veri senkronizasyonu (Ürünler + Geçmiş)."""

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.setWindowTitle("Firebase Bulut Senkronizasyon")
        self.setMinimumWidth(450)


        self.json_dosya_adi = FIREBASE_KEY_PATH
        self.database_url = FIREBASE_DB_URL


        layout = QVBoxLayout(self)

        if not FIREBASE_MUMKUN:
            layout.addWidget(
                QLabel("HATA: 'firebase-admin' yüklü değil.\n'pip install firebase-admin' komutunu çalıştırın."))
            return

        baslik = QLabel("Firebase Veri Yedekleme")
        baslik.setStyleSheet("font-size: 16px; font-weight: bold; color: #ffa000;")
        baslik.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(baslik)

        info = QLabel("Ürünlerinizi ve satış geçmişinizi bulutla eşitleyin.")
        info.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(info)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setStyleSheet("QProgressBar::chunk { background-color: #ffa000; }")
        layout.addWidget(self.progress)

        self.status_lbl = QLabel("Hazır...")
        self.status_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.status_lbl)

        btn_layout = QHBoxLayout()
        self.btn_gonder = QPushButton("Buluta Yükle (Yedekle)")
        self.btn_gonder.clicked.connect(self.buluta_gonder)

        self.btn_cek = QPushButton("Buluttan İndir (Geri Yükle)")
        self.btn_cek.clicked.connect(self.buluttan_cek)

        btn_layout.addWidget(self.btn_gonder)
        btn_layout.addWidget(self.btn_cek)
        layout.addLayout(btn_layout)

    def baglanti_kur(self):
        """
        Firebase projesine authentication (kimlik doğrulama) yapar.
        'firebase_key.json' dosyasını okur ve bağlantı nesnesini oluşturur.
        """
        if not firebase_admin._apps:
            try:
                cred = credentials.Certificate(self.json_dosya_adi)
                firebase_admin.initialize_app(cred, {
                    'databaseURL': self.database_url
                })
            except FileNotFoundError:
                QMessageBox.critical(self, "Hata",
                                     f"'{self.json_dosya_adi}' dosyası bulunamadı!")
                return False
            except Exception as e:
                QMessageBox.critical(self, "Bağlantı Hatası", str(e))
                return False
        return True

    def buluta_gonder(self):
        """
        Yerel veritabanındaki tüm verileri (Ürünler + Geçmiş + Kullanıcılar)
        JSON formatına çevirip Firebase'e yükler.
        """
        if not self.baglanti_kur(): return

        self.status_lbl.setText("Veriler hazırlanıyor...")
        self.progress.setValue(10)
        QApplication.processEvents()

        try:
            # 1. ÜRÜNLERİ HAZIRLA
            urunler = self.veritabani.urunleri_getir()
            urunler_export = {}
            for urun in urunler:
                u_id = str(urun[0])
                urunler_export[u_id] = {
                    "urun_kodu": urun[1], "ad": urun[2], "kategori": urun[3],
                    "fiyat": urun[4], "miktar": urun[5], "birim": urun[6],
                    "min_stok": urun[7], "baslangic_miktari": urun[8],
                    "son_kullanma_tarihi": urun[9]
                }

            # 2. STOK HAREKETLERİNİ HAZIRLA
            self.veritabani.cursor.execute("SELECT * FROM stok_hareketleri")
            hareketler = self.veritabani.cursor.fetchall()
            hareketler_export = {}
            for h in hareketler:
                h_id = str(h[0])
                tarih_str = str(h[6])
                hareketler_export[h_id] = {
                    "urun_id": h[1], "kullanici_adi": h[2], "islem_tipi": h[3],
                    "miktar_degisimi": h[4], "yeni_miktar": h[5], "tarih": tarih_str,
                    "notlar": h[7], "satis_fiyati": h[8]
                }

            #  3. KULLANICILARI HAZIRLA
            kullanicilar = self.veritabani.tum_kullanicilari_getir()
            kullanicilar_export = {}
            for k in kullanicilar:
                k_id = str(k[0])
                kullanicilar_export[k_id] = {
                    "kullanici_adi": k[1],
                    "sifre_hash": k[2]  # Şifreyi hashli haliyle yedekliyoruz
                }


            full_data = {
                "urunler": urunler_export,
                "hareketler": hareketler_export,
                "kullanicilar": kullanicilar_export  # Listeye ekledik
            }

            self.status_lbl.setText("Firebase'e yükleniyor...")
            self.progress.setValue(50)
            QApplication.processEvents()

            ref = db.reference('tam_yedek')
            ref.set(full_data)

            self.progress.setValue(100)
            self.status_lbl.setText("Yedekleme Başarılı!")
            QMessageBox.information(self, "Başarılı", "Ürünler, Geçmiş ve Kullanıcılar buluta yüklendi.")

        except Exception as e:
            self.status_lbl.setText("Hata oluştu.")
            error_str = str(e)
            if "Invalid JWT Signature" in error_str or "invalid_grant" in error_str:
                QMessageBox.critical(self, "Bağlantı Hatası (Yetkilendirme)",
                                     "Bulut sunucusu bağlantıyı reddetti.\n\n"
                                     "Olası Sebepler:\n"
                                     "1. Bilgisayar saatiniz sunucu saatiyle uyuşmuyor (En sık görülen).\n"
                                     "2. 'firebase_key.json' dosyası hatalı veya eski.\n\n"
                                     "Lütfen Windows saat ayarlarınızın 'Otomatik' olduğundan emin olun.")
            else:
                QMessageBox.critical(self, "Hata", f"Yedekleme hatası: {error_str}")

    def buluttan_cek(self):
        """
        Firebase üzerindeki yedekten verileri indirir ve yerel veritabanını günceller.
        DİKKAT: Mevcut yerel verileri tamamen siler!
        """
        if not self.baglanti_kur(): return

        msg = QMessageBox(self)
        msg.setWindowTitle("DİKKAT")
        msg.setText(
            "Buluttan indirmek, MEVCUT VERİLERİ (Kullanıcılar dahil) SİLİP üzerine yazacaktır.\nDevam etmek istiyor musunuz?")
        msg.setIcon(QMessageBox.Icon.Warning)
        btn_evet = msg.addButton("Evet", QMessageBox.ButtonRole.YesRole)
        btn_hayir = msg.addButton("Hayır", QMessageBox.ButtonRole.NoRole)
        msg.setDefaultButton(btn_hayir)
        msg.exec()

        if msg.clickedButton() != btn_evet:
            return

        self.status_lbl.setText("Veriler indiriliyor...")
        self.progress.setValue(20)
        QApplication.processEvents()

        try:
            ref = db.reference('tam_yedek')
            snapshot = ref.get()

            if not snapshot:
                QMessageBox.warning(self, "Uyarı", "Bulutta yedek bulunamadı.")
                return

            self.status_lbl.setText("Veritabanına işleniyor...")
            self.progress.setValue(60)
            QApplication.processEvents()

            cursor = self.veritabani.baglanti.cursor()

            # Önce tabloları temizle
            cursor.execute("DELETE FROM stok_hareketleri")
            cursor.execute("DELETE FROM urunler")
            cursor.execute("DELETE FROM kullanicilar")


            # 1. ÜRÜNLERİ GERİ YÜKLE
            urunler_data = snapshot.get('urunler', {})
            iterable_urun = enumerate(urunler_data) if isinstance(urunler_data, list) else urunler_data.items()

            for key, val in iterable_urun:
                if val is None: continue
                u_id = int(key) if isinstance(urunler_data, dict) else val.get('id', key)
                cursor.execute("""
                    INSERT INTO urunler (id, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, baslangic_miktari, son_kullanma_tarihi)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    u_id, val.get('urun_kodu'), val.get('ad'), val.get('kategori'),
                    float(val.get('fiyat', 0)), float(val.get('miktar', 0)), val.get('birim', 'adet'),
                    float(val.get('min_stok', 0)), float(val.get('baslangic_miktari', 0)),
                    val.get('son_kullanma_tarihi')
                ))

            # 2. HAREKETLERİ GERİ YÜKLE
            hareket_data = snapshot.get('hareketler', {})
            iterable_har = enumerate(hareket_data) if isinstance(hareket_data, list) else hareket_data.items()

            for key, val in iterable_har:
                if val is None: continue
                h_id = int(key) if isinstance(hareket_data, dict) else val.get('id', key)
                cursor.execute("""
                    INSERT INTO stok_hareketleri (id, urun_id, kullanici_adi, islem_tipi, miktar_degisimi, yeni_miktar, tarih, notlar, satis_fiyati)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    h_id, val.get('urun_id'), val.get('kullanici_adi'), val.get('islem_tipi'),
                    float(val.get('miktar_degisimi', 0)), float(val.get('yeni_miktar', 0)),
                    val.get('tarih'), val.get('notlar'), val.get('satis_fiyati')
                ))

            # 3. KULLANICILARI GERİ YÜKLE
            kullanici_data = snapshot.get('kullanicilar', {})
            iterable_kul = enumerate(kullanici_data) if isinstance(kullanici_data, list) else kullanici_data.items()

            for key, val in iterable_kul:
                if val is None: continue
                k_id = int(key) if isinstance(kullanici_data, dict) else val.get('id', key)
                # Yeni yazdığımız ham veri yükleme metodunu kullanıyoruz
                self.veritabani.kullanici_yukle_raw(k_id, val.get('kullanici_adi'), val.get('sifre_hash'))


            self.veritabani.baglanti.commit()
            self.progress.setValue(100)
            self.status_lbl.setText("Tamamlandı!")
            QMessageBox.information(self, "Başarılı", "Tüm veriler (Kullanıcılar dahil) başarıyla geri yüklendi.")

        except Exception as e:
            self.veritabani.baglanti.rollback()
            self.status_lbl.setText("Hata!")
            QMessageBox.critical(self, "Hata", str(e))


class SatisDialog(QDialog):
    """
    Hızlı satış işlemi için geliştirilmiş pop-up penceresidir.
    Kullanıcıdan satış miktarını ve birim fiyatını alır.
    """

    def __init__(self, urun_adi, mevcut_stok, birim, guncel_alis_fiyati, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Satış Yap")
        self.setModal(True)
        self.setMinimumWidth(350)

        layout = QFormLayout(self)

        self.lbl_info = QLabel(f"<b>{urun_adi}</b><br>Mevcut Stok: {mevcut_stok:g} {birim}")
        self.lbl_info.setStyleSheet("color: #94a3b8; font-size: 13px; margin-bottom: 10px;")
        layout.addRow(self.lbl_info)

        self.lbl_alis = QLabel(f"Sistemdeki Alış/Maliyet Fiyatı: <b>{guncel_alis_fiyati:.2f}₺</b>")
        self.lbl_alis.setStyleSheet("color: #64748b; font-size: 11px;")
        layout.addRow(self.lbl_alis)

        self.input_miktar = QLineEdit()
        self.input_miktar.setPlaceholderText("Miktar girin...")
        self.input_miktar.setValidator(QDoubleValidator(0.0, 999999.0, 2))
        layout.addRow("Satılacak Miktar:", self.input_miktar)

        self.input_fiyat = QLineEdit(str(guncel_alis_fiyati))
        self.input_fiyat.setPlaceholderText("Birim satış fiyatı...")
        self.input_fiyat.setValidator(QDoubleValidator(0.0, 999999.0, 2))
        layout.addRow("Birim Satış Fiyatı (₺):", self.input_fiyat)

        # --- BUTONLARI TÜRKÇELEŞTİRME KISMI ---
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)

        # Standart butonların metinlerini değiştiriyoruz
        self.buttonBox.button(QDialogButtonBox.StandardButton.Ok).setText("Tamam")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel).setText("İptal Et")
        # --------------------------------------

        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        layout.addRow(self.buttonBox)

    def get_values(self):
        """Kullanıcının girdiği miktar ve fiyat bilgisini döner."""
        try:
            miktar_text = self.input_miktar.text().replace(',', '.')
            fiyat_text = self.input_fiyat.text().replace(',', '.')
            if not miktar_text: return None, None
            m = float(miktar_text)
            f = float(fiyat_text) if fiyat_text else 0.0
            return m, f
        except ValueError:
            return None, None


class UrunDuzenlemeDialog(QDialog):
    """
    Mevcut bir ürünün detaylarını (Ad, Kategori, Fiyat vb.) düzenlemek için kullanılır.
    Stok miktarı buradan değiştirilmez (güvenlik ve log takibi için), sadece kart bilgileri güncellenir.
    """

    def __init__(self, urun_detaylari, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Ürün Bilgilerini Düzenle")
        self.form_layout = QFormLayout(self)
        self.urun_id = urun_detaylari[0]
        self.urun_kodu_label = QLabel(urun_detaylari[1])
        self.ad_input = QLineEdit(urun_detaylari[2])
        self.kategori_input = QLineEdit(urun_detaylari[3])
        self.fiyat_input = QLineEdit(str(urun_detaylari[4]))
        self.yeni_birim_input = QComboBox()
        self.yeni_birim_input.addItems(STANDART_BIRIMLER)
        self.yeni_birim_input.setFixedWidth(100)
        self.yeni_birim_input.setEditable(True)
        self.yeni_birim_input.lineEdit().setReadOnly(True)
        self.yeni_birim_input.lineEdit().setAlignment(Qt.AlignmentFlag.AlignCenter)
        mevcut_birim = urun_detaylari[6]
        index = self.yeni_birim_input.findText(mevcut_birim)
        if index >= 0: self.yeni_birim_input.setCurrentIndex(index)
        self.min_stok_input = QLineEdit(str(urun_detaylari[7]))
        self.skt_input = QDateEdit()
        self.skt_input.setCalendarPopup(True)
        self.skt_input.setDisplayFormat("yyyy-MM-dd")
        mevcut_skt_str = urun_detaylari[8]
        if mevcut_skt_str:
            tarih = QDate.fromString(mevcut_skt_str, "yyyy-MM-dd")
            self.skt_input.setDate(tarih)
        else:
            self.skt_input.setDate(QDate.currentDate().addYears(1))
        self.form_layout.addRow("Ürün ID (Sistem):", QLabel(str(self.urun_id)))
        self.form_layout.addRow("Ürün Kodu (SKU):", self.urun_kodu_label)
        self.form_layout.addRow("Ürün Adı:", self.ad_input)
        self.form_layout.addRow("Kategori:", self.kategori_input)
        self.form_layout.addRow("Alış Fiyatı (Maliyet ₺):", self.fiyat_input)
        self.form_layout.addRow("Birim:", self.yeni_birim_input)
        self.form_layout.addRow("Min. Stok:", self.min_stok_input)
        self.form_layout.addRow("Son Kul. Tarihi:", self.skt_input)

        float_validator = QDoubleValidator()
        float_validator.setBottom(0.0)
        self.fiyat_input.setValidator(float_validator)
        self.min_stok_input.setValidator(float_validator)
        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel).setText("İptal Et")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Ok).setText("Kaydet")
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        self.form_layout.addWidget(self.buttonBox)

    def get_data(self):
        """Formdaki güncellenmiş verileri doğrular ve tuple olarak döner."""
        try:
            fiyat = float(self.fiyat_input.text().replace(',', '.'))
            min_stok = float(self.min_stok_input.text().replace(',', '.'))
            birim = self.yeni_birim_input.currentText()
            ad = self.ad_input.text().strip().upper()
            kategori = self.kategori_input.text().strip().upper()
            skt = self.skt_input.date().toString("yyyy-MM-dd")
            if not ad:
                QMessageBox.warning(self, "Hata", "Ürün adı boş bırakılamaz.")
                return None
            return self.urun_id, ad, kategori, fiyat, birim, min_stok, skt
        except ValueError:
            QMessageBox.warning(self, "Hata", "Fiyat ve Min. Stok sayısal değer olmalıdır.")
            return None


class YeniKullaniciDialog(QDialog):
    """
    Sisteme yeni yönetici/personel eklemek için kullanılan kayıt formu.
    Şifre eşleşme kontrolü yapar ve veritabanına hashing ile kaydeder.
    """

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.setWindowTitle("Yeni Kullanıcı Oluştur")
        self.setObjectName("authWindow")
        self.setMinimumWidth(400)

        if parent:
            # Eğer ana pencere üzerinden açıldıysa, ANA PENCERENİN ortasında açıl
            parent_geo = parent.geometry()
            self_geo = self.geometry()
            x = parent_geo.x() + (parent_geo.width() - self.width()) // 2
            y = parent_geo.y() + (parent_geo.height() - self.height()) // 2
            self.move(x, y)
        else:
            # Eğer ana pencere yoksa, EKRANIN ortasında açıl
            screen = QApplication.primaryScreen()
            if screen:
                rect = screen.availableGeometry()
                center = rect.center()
                frame_geo = self.frameGeometry()
                frame_geo.moveCenter(center)
                self.move(frame_geo.topLeft())

        self.ana_layout = QVBoxLayout(self)

        # --- LOGO AYARLARI ---
        logo_layout = QVBoxLayout()
        logo_layout.setContentsMargins(0, 0, 0, 20)

        logo_icon = QLabel()
        logo_icon.setObjectName("logoIcon")
        try:
            logo_path = dosya_yolunu_bul("StockFlow_Logo.png")
            logo_pixmap = QPixmap(logo_path)
            if not logo_pixmap.isNull():
                logo_icon.setPixmap(logo_pixmap.scaled(400, 150, Qt.AspectRatioMode.KeepAspectRatio,
                                                       Qt.TransformationMode.SmoothTransformation))
                logo_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
            else:
                raise FileNotFoundError
        except FileNotFoundError:
            # Logo yoksa boş kalsın veya basit metin
            logo_icon.setText("STOCKFLOW")
            logo_icon.setStyleSheet(
                "font-size: 24px; font-weight: bold; color: #3b82f6; qproperty-alignment: AlignCenter;")

        logo_text = QLabel("Yeni Hesap Oluştur")
        logo_text.setObjectName("logoText")
        logo_text.setStyleSheet("font-size: 20px; font-weight: bold; qproperty-alignment: AlignCenter;")

        logo_layout.addWidget(logo_icon)
        logo_layout.addWidget(logo_text)
        self.ana_layout.addLayout(logo_layout)

        # --- FORM KISMI ---
        layout = QFormLayout()
        layout.setContentsMargins(20, 0, 20, 0)
        layout.setVerticalSpacing(15)
        self.k_adi = QLineEdit()
        self.k_adi.setPlaceholderText("Yeni Kullanıcı Adı")
        self.sifre = QLineEdit()
        self.sifre.setPlaceholderText("Şifre")
        self.sifre.setEchoMode(QLineEdit.EchoMode.Password)
        self.sifre_t = QLineEdit()
        self.sifre_t.setPlaceholderText("Şifre Tekrar")
        self.sifre_t.setEchoMode(QLineEdit.EchoMode.Password)
        layout.addRow("Kullanıcı Adı:", self.k_adi)
        layout.addRow("Şifre:", self.sifre)
        layout.addRow("Şifre Tekrar:", self.sifre_t)
        self.ana_layout.addLayout(layout)

        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        self.buttonBox.button(QDialogButtonBox.StandardButton.Ok).setText("Oluştur")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel).setText("Geri")
        self.buttonBox.accepted.connect(self.hesap_olustur)
        self.buttonBox.rejected.connect(self.reject)
        self.ana_layout.addWidget(self.buttonBox)
        self.sifre_t.returnPressed.connect(self.hesap_olustur)

    def hesap_olustur(self):
        k_adi = self.k_adi.text().strip()
        sifre = self.sifre.text()
        sifre_t = self.sifre_t.text()
        if sifre != sifre_t:
            QMessageBox.warning(self, "Hata", "Şifreler eşleşmiyor.")
            return
        basari, mesaj = self.veritabani.kullanici_ekle(k_adi, sifre)
        if basari:
            QMessageBox.information(self, "Başarılı",
                                    f"'{k_adi}' kullanıcısı başarıyla oluşturuldu. Şimdi giriş yapabilirsiniz.")
            self.accept()
        else:
            QMessageBox.warning(self, "Hata", mesaj)


class FiltreDialog(QDialog):
    """
    Ana stok listesi için gelişmiş filtreleme seçenekleri sunar.
    Fiyat aralığı, stok aralığı ve sadece düşük stokluları gösterme seçenekleri içerir.
    """

    def __init__(self, mevcut_filtreler=None, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gelişmiş Filtrele")
        self.setMinimumWidth(400)
        layout = QFormLayout(self)
        if mevcut_filtreler is None: mevcut_filtreler = {}
        self.min_fiyat = QLineEdit(str(mevcut_filtreler.get("min_fiyat", "")))
        self.max_fiyat = QLineEdit(str(mevcut_filtreler.get("max_fiyat", "")))
        fiyat_layout = QHBoxLayout()
        fiyat_layout.addWidget(self.min_fiyat)
        fiyat_layout.addWidget(QLabel("-"))
        fiyat_layout.addWidget(self.max_fiyat)
        layout.addRow("Fiyat Aralığı (₺):", fiyat_layout)
        self.min_stok = QLineEdit(str(mevcut_filtreler.get("min_stok", "")))
        self.max_stok = QLineEdit(str(mevcut_filtreler.get("max_stok", "")))
        stok_layout = QHBoxLayout()
        stok_layout.addWidget(self.min_stok)
        stok_layout.addWidget(QLabel("-"))
        stok_layout.addWidget(self.max_stok)
        layout.addRow("Stok Aralığı:", stok_layout)
        float_validator = QDoubleValidator()
        float_validator.setBottom(0.0)
        self.min_fiyat.setValidator(float_validator)
        self.max_fiyat.setValidator(float_validator)
        self.min_stok.setValidator(float_validator)
        self.max_stok.setValidator(float_validator)
        self.sadece_dusuk_stok = QCheckBox("Sadece düşük stoktakileri göster")
        self.sadece_dusuk_stok.setChecked(mevcut_filtreler.get("dusuk_stok_only", False))
        layout.addRow("", self.sadece_dusuk_stok)
        self.buttonBox = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel | QDialogButtonBox.StandardButton.Reset)
        self.buttonBox.button(QDialogButtonBox.StandardButton.Ok).setText("Filtrele")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Reset).setText("Filtreyi Temizle")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel).setText("İptal Et")
        self.buttonBox.accepted.connect(self.accept)
        self.buttonBox.rejected.connect(self.reject)
        self.buttonBox.button(QDialogButtonBox.StandardButton.Reset).clicked.connect(self.filtreyi_temizle_ve_kapat)
        layout.addWidget(self.buttonBox)
        self.reset_requested = False

    def filtreyi_temizle_ve_kapat(self):
        """Tüm seçimleri sıfırlar ve filtreyi kaldırır."""
        self.reset_requested = True
        self.accept()

    def get_filtreler(self):
        """Seçilen filtre kriterlerini bir sözlük olarak döndürür."""

        def to_float(widget):
            try:
                return float(widget.text().replace(',', '.'))
            except ValueError:
                return None

        return {
            "min_fiyat": to_float(self.min_fiyat), "max_fiyat": to_float(self.max_fiyat),
            "min_stok": to_float(self.min_stok), "max_stok": to_float(self.max_stok),
            "dusuk_stok_only": self.sadece_dusuk_stok.isChecked()
        }


# =============================================================================
# 4. UI WIDGETS AND PAGES
# =============================================================================

class KarZararSayfasi(QWidget):
    """
    İşletmenin kar/zarar durumunu analiz eden raporlama sayfasıdır.
    Satışlardan elde edilen ciroyu ve maliyetleri karşılaştırarak net karı hesaplar.
    """

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.arayuz_olustur()
        self.raporu_guncelle()

    def create_metric_card(self, parent_layout, title, value_id):
        """Sayfanın üst kısmındaki özet bilgi kartlarını (Toplam Ciro vb.) oluşturur."""
        card_frame = QFrame()
        card_frame.setObjectName("metricCard")
        card_layout = QVBoxLayout(card_frame)
        card_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label = QLabel(title)
        title_label.setObjectName("metricTitle")
        value_label = QLabel("0.00₺")
        value_label.setObjectName("metricValue")
        if title == "TOPLAM KÂR": value_label.setStyleSheet("color: #4ade80;")
        # card_layout.addWidget(icon_label) # Icon removed
        card_layout.addWidget(title_label)
        card_layout.addWidget(value_label)
        parent_layout.addWidget(card_frame)
        if not hasattr(self, 'metric_labels'): self.metric_labels = {}
        self.metric_labels[value_id] = value_label

    def arayuz_olustur(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)
        ust_panel = QHBoxLayout()
        ust_panel.setSpacing(20)
        filtre_frame = QFrame()
        filtre_layout = QVBoxLayout(filtre_frame)
        filtre_layout.addWidget(QLabel("Zaman Aralığı:"))
        self.aralik_secimi = QComboBox()
        self.aralik_secimi.addItem(" Bu Hafta", "haftalik")
        self.aralik_secimi.addItem(" Bu Ay", "aylik")
        self.aralik_secimi.addItem(" Bu Yıl", "yillik")
        self.aralik_secimi.addItem(" Tüm Zamanlar", "tumu")
        self.aralik_secimi.setCurrentIndex(1)
        self.aralik_secimi.currentIndexChanged.connect(self.raporu_guncelle)
        filtre_layout.addWidget(self.aralik_secimi)
        yenile_btn = QPushButton("Hesapla")
        yenile_btn.clicked.connect(self.raporu_guncelle)
        filtre_layout.addWidget(yenile_btn)
        ust_panel.addWidget(filtre_frame)
        self.create_metric_card(ust_panel, "SATIŞ ADEDİ", "toplam_adet")
        self.create_metric_card(ust_panel, "TOPLAM CİRO", "toplam_ciro")
        self.create_metric_card(ust_panel, "TOPLAM KÂR", "toplam_kar")
        layout.addLayout(ust_panel)
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(7)
        self.tablo.setHorizontalHeaderLabels(
            ["Ürün Kodu", "Ürün Adı", "Kategori", "Adet", "Alış Fiyatı", "Satış Fiyatı", "Kâr/Zarar"])
        header = self.tablo.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tablo.setSortingEnabled(True)
        self.tablo.verticalHeader().setVisible(False)
        layout.addWidget(self.tablo)
        layout.addWidget(
            QLabel("Kâr Hesabı: (Satış Anındaki Fiyat - Güncel Alış Fiyatı) x Adet formülü ile hesaplanır."))

    def raporu_guncelle(self):
        """Seçilen zaman aralığına göre veritabanından verileri çeker ve tabloyu yeniler."""
        self.tablo.setSortingEnabled(False)
        self.tablo.setRowCount(0)
        aralik = self.aralik_secimi.currentData()
        veriler = self.veritabani.kar_zarar_raporu_getir(aralik)
        genel_toplam_kar = 0.0
        genel_toplam_ciro = 0.0
        genel_satis_adedi = 0
        self.tablo.setRowCount(len(veriler))
        for satir, (kod, ad, kategori, satis_adedi, alis_fiyati, ort_satis_fiyati, urun_ciro, urun_kar) in enumerate(
                veriler):
            genel_toplam_kar += urun_kar
            genel_toplam_ciro += urun_ciro
            genel_satis_adedi += satis_adedi
            self.tablo.setItem(satir, 0, QTableWidgetItem(kod))
            self.tablo.setItem(satir, 1, QTableWidgetItem(ad))
            self.tablo.setItem(satir, 2, QTableWidgetItem(kategori))
            self.tablo.setItem(satir, 3, NumericTableWidgetItem(str(satis_adedi), satis_adedi))
            self.tablo.setItem(satir, 4, NumericTableWidgetItem(f"{alis_fiyati:.2f}₺", alis_fiyati))
            self.tablo.setItem(satir, 5, NumericTableWidgetItem(f"{ort_satis_fiyati:.2f}₺", ort_satis_fiyati))
            kar_item = NumericTableWidgetItem(f"{urun_kar:.2f}₺", urun_kar)
            if urun_kar > 0:
                kar_item.setForeground(QColor("#4ade80"))
            elif urun_kar < 0:
                kar_item.setForeground(QColor("#f87171"))
            self.tablo.setItem(satir, 6, kar_item)
        self.metric_labels["toplam_adet"].setText(str(int(genel_satis_adedi)))
        self.metric_labels["toplam_ciro"].setText(f"{genel_toplam_ciro:,.2f}₺")
        self.metric_labels["toplam_kar"].setText(f"{genel_toplam_kar:,.2f}₺")
        self.tablo.resizeRowsToContents()
        self.tablo.setSortingEnabled(True)


class SatisRaporuSayfasi(QWidget):
    """
    En çok satan ürünleri listeleyen performans raporu sayfasıdır.
    Hangi ürünün ne kadar sattığını adet ve ciro bazında gösterir.
    """

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.arayuz_olustur()
        self.raporu_guncelle()

    def arayuz_olustur(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(25, 25, 25, 25)
        layout.setSpacing(20)
        ust_panel = QHBoxLayout()
        ust_panel.addWidget(QLabel("Rapor Aralığı:"))
        self.aralik_secimi = QComboBox()
        self.aralik_secimi.addItem(" Bu Hafta", "haftalik")
        self.aralik_secimi.addItem(" Bu Ay", "aylik")
        self.aralik_secimi.addItem(" Bu Yıl", "yillik")
        self.aralik_secimi.setCurrentIndex(1)
        self.aralik_secimi.currentIndexChanged.connect(self.raporu_guncelle)
        ust_panel.addWidget(self.aralik_secimi)
        yenile_btn = QPushButton("Yenile")
        yenile_btn.clicked.connect(self.raporu_guncelle)
        ust_panel.addStretch()
        ust_panel.addWidget(yenile_btn)
        layout.addLayout(ust_panel)
        self.tablo = QTableWidget()
        self.tablo.setColumnCount(5)
        self.tablo.setHorizontalHeaderLabels(["Sıra", "Ürün Kodu", "Ürün Adı", "Kategori", "Toplam Satış"])
        header = self.tablo.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.tablo.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tablo.verticalHeader().setVisible(False)
        layout.addWidget(self.tablo)
        info_lbl = QLabel("Not: Bu rapor sadece 'Stok Çıkışı' işlemlerini satış olarak kabul eder.")
        info_lbl.setStyleSheet("color: #64748b; font-style: italic;")
        layout.addWidget(info_lbl)

    def raporu_guncelle(self):
        self.tablo.setRowCount(0)
        aralik = self.aralik_secimi.currentData()
        veriler = self.veritabani.en_cok_satanlari_getir(aralik)
        self.tablo.setRowCount(len(veriler))
        for satir, (kod, ad, kategori, toplam_satis, birim) in enumerate(veriler):
            sira_item = QTableWidgetItem(str(satir + 1))
            sira_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            kod_item = QTableWidgetItem(kod)
            kod_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            ad_item = QTableWidgetItem(ad)
            kat_item = QTableWidgetItem(kategori)
            satis_str = f"{toplam_satis:g} {birim}"
            satis_item = NumericTableWidgetItem(satis_str, toplam_satis)
            satis_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            if toplam_satis > 0:
                satis_item.setForeground(QColor("#4ade80"))
                satis_item.setBackground(QColor("#1c3d34"))
            else:
                satis_item.setForeground(QColor("#94a3b8"))
            self.tablo.setItem(satir, 0, sira_item)
            self.tablo.setItem(satir, 1, kod_item)
            self.tablo.setItem(satir, 2, ad_item)
            self.tablo.setItem(satir, 3, kat_item)
            self.tablo.setItem(satir, 4, satis_item)


class DusukStokSayfasi(QWidget):
    """
    Kritik stok seviyesinin altına düşen ürünleri acil sipariş listesi şeklinde gösterir.
    Stok takibi için hayati öneme sahiptir.
    """

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.arayuz_olustur()
        self.stogu_guncelle()

    def arayuz_olustur(self):
        ana_duzen = QVBoxLayout(self)
        ana_duzen.setContentsMargins(25, 25, 25, 25)
        ana_duzen.setSpacing(20)
        yenile_btn = QPushButton("Listeyi Yenile")
        yenile_btn.clicked.connect(self.stogu_guncelle)
        yenile_btn.setFixedWidth(200)
        ana_duzen.addWidget(yenile_btn, alignment=Qt.AlignmentFlag.AlignLeft)
        self.stok_tablosu = QTableWidget()
        self.stok_tablosu.setObjectName("stokTablosu")
        self.stok_tablosu.setColumnCount(7)
        self.stok_tablosu.setHorizontalHeaderLabels(
            ["ID", "Ürün Kodu", "Ürün Adı", "Kategori", "Mevcut Miktar", "Birim", "Min. Stok"])
        header = self.stok_tablosu.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.stok_tablosu.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stok_tablosu.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stok_tablosu.setSortingEnabled(True)
        self.stok_tablosu.sortByColumn(2, Qt.SortOrder.AscendingOrder)
        self.stok_tablosu.verticalHeader().setVisible(False)
        ana_duzen.addWidget(self.stok_tablosu, 1)

    def stogu_guncelle(self):
        self.stok_tablosu.setSortingEnabled(False)
        self.stok_tablosu.setRowCount(0)
        urun_listesi = self.veritabani.dusuk_stok_urunleri_getir()
        self.stok_tablosu.setRowCount(len(urun_listesi))
        for satir, (id_val, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok) in enumerate(urun_listesi):
            id_item = NumericTableWidgetItem(str(id_val), id_val)
            kod_item = QTableWidgetItem(urun_kodu)
            ad_item = QTableWidgetItem(ad)
            kategori_item = QTableWidgetItem(kategori)
            miktar_str = f"{miktar:g}"
            min_stok_str = f"{min_stok:g}"
            miktar_item = NumericTableWidgetItem(miktar_str, miktar)
            birim_item = QTableWidgetItem(birim)
            min_stok_item = NumericTableWidgetItem(min_stok_str, min_stok)
            id_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            kod_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            birim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            min_stok_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.stok_tablosu.setItem(satir, 0, id_item)
            self.stok_tablosu.setItem(satir, 1, kod_item)
            self.stok_tablosu.setItem(satir, 2, ad_item)
            self.stok_tablosu.setItem(satir, 3, kategori_item)
            self.stok_tablosu.setItem(satir, 4, miktar_item)
            self.stok_tablosu.setItem(satir, 5, birim_item)
            self.stok_tablosu.setItem(satir, 6, min_stok_item)
            for col in range(self.stok_tablosu.columnCount()):
                item = self.stok_tablosu.item(satir, col)
                if item: item.setBackground(QColor("#4a1f2c"))
            miktar_item.setText(f"{miktar_str}")
            miktar_item.setForeground(QColor("#f87171"))
        self.stok_tablosu.resizeRowsToContents()
        self.stok_tablosu.setSortingEnabled(True)


class StokHareketSayfasi(QWidget):
    """
    Tüm stok giriş-çıkış işlemlerinin (Log) tutulduğu ve görüntülendiği sayfadır.
    Hatalı işlemler buradan geri alınabilir (Undo).
    """

    def __init__(self, veritabani_yoneticisi, kullanici_adi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.kullanici_adi = kullanici_adi
        self.arayuz_olustur()
        self.raporu_guncelle()

    def arayuz_olustur(self):
        ana_duzen = QVBoxLayout(self)
        ana_duzen.setContentsMargins(25, 25, 25, 25)
        ana_duzen.setSpacing(20)
        ust_bar = QHBoxLayout()
        ust_bar.addWidget(QLabel("Zaman Aralığı:"))
        self.zaman_filtresi = QComboBox()
        self.zaman_filtresi.addItem("Tüm Zamanlar", "tumu")
        self.zaman_filtresi.addItem("Son 30 Gün", "aylik")
        self.zaman_filtresi.addItem("Son 7 Gün", "haftalik")
        self.zaman_filtresi.setFixedWidth(150)
        self.zaman_filtresi.currentIndexChanged.connect(self.raporu_guncelle)
        ust_bar.addWidget(self.zaman_filtresi)
        yenile_btn = QPushButton("Raporu Yenile")
        yenile_btn.setFixedWidth(200)
        yenile_btn.clicked.connect(self.raporu_guncelle)
        ust_bar.addStretch()
        ust_bar.addWidget(yenile_btn)
        ana_duzen.addLayout(ust_bar)
        self.rapor_tablosu = QTableWidget()
        self.rapor_tablosu.setObjectName("stokTablosu")
        self.rapor_tablosu.setColumnCount(10)
        self.rapor_tablosu.setHorizontalHeaderLabels(
            ["Tarih", "Kullanıcı", "Ürün Kodu", "Ürün Adı", "İşlem", "Değişim", "Yeni Miktar", "Birim", "İşlem Fiyatı",
             "Notlar"])
        header = self.rapor_tablosu.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.Stretch)
        self.rapor_tablosu.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.rapor_tablosu.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.rapor_tablosu.setSortingEnabled(True)
        self.rapor_tablosu.sortByColumn(0, Qt.SortOrder.DescendingOrder)
        self.rapor_tablosu.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.rapor_tablosu.customContextMenuRequested.connect(self.menu_goster)
        self.rapor_tablosu.verticalHeader().setVisible(False)
        ana_duzen.addWidget(self.rapor_tablosu, 1)
        info_label = QLabel("Bir işlemi geri almak için satıra sağ tıklayın.")
        info_label.setStyleSheet("color: #64748b; font-style: italic;")
        ana_duzen.addWidget(info_label)

    def raporu_guncelle(self):
        self.rapor_tablosu.setSortingEnabled(False)
        self.rapor_tablosu.setRowCount(0)
        zaman_araligi = self.zaman_filtresi.currentData()
        hareket_listesi = self.veritabani.stok_hareketlerini_getir(zaman_araligi)
        self.rapor_tablosu.setRowCount(len(hareket_listesi))
        for satir, (hareket_id, tarih, kullanici, kod, ad, islem, degisim, yeni_miktar, birim, notlar,
                    satis_fiyati) in enumerate(hareket_listesi):
            tarih_str = tarih.strftime("%Y-%m-%d %H:%M") if isinstance(tarih, datetime) else str(tarih)
            tarih_item = QTableWidgetItem(tarih_str)
            tarih_item.setData(Qt.ItemDataRole.UserRole, hareket_id)
            kullanici_item = QTableWidgetItem(kullanici)
            kod_item = QTableWidgetItem(kod if kod else "N/A")
            ad_item = QTableWidgetItem(ad if ad else "SİLİNMİŞ ÜRÜN")
            islem_item = QTableWidgetItem(islem)
            degisim_str = f"{degisim:+g}"
            yeni_miktar_str = f"{yeni_miktar:g}" if yeni_miktar is not None else "N/A"
            fiyat_str = f"{satis_fiyati:g}₺" if satis_fiyati is not None else "-"
            degisim_item = NumericTableWidgetItem(degisim_str, degisim)
            yeni_miktar_item = NumericTableWidgetItem(yeni_miktar_str, yeni_miktar or 0)
            birim_item = QTableWidgetItem(birim if birim else "N/A")
            fiyat_item = QTableWidgetItem(fiyat_str)
            notlar_item = QTableWidgetItem(notlar)
            degisim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            yeni_miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            islem_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            birim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            fiyat_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.rapor_tablosu.setItem(satir, 0, tarih_item)
            self.rapor_tablosu.setItem(satir, 1, kullanici_item)
            self.rapor_tablosu.setItem(satir, 2, kod_item)
            self.rapor_tablosu.setItem(satir, 3, ad_item)
            self.rapor_tablosu.setItem(satir, 4, islem_item)
            self.rapor_tablosu.setItem(satir, 5, degisim_item)
            self.rapor_tablosu.setItem(satir, 6, yeni_miktar_item)
            self.rapor_tablosu.setItem(satir, 7, birim_item)
            self.rapor_tablosu.setItem(satir, 8, fiyat_item)
            self.rapor_tablosu.setItem(satir, 9, notlar_item)
            if islem == "STOK EKLEME" or islem == "YENİ ÜRÜN" or islem == "İŞLEM GERİ ALMA":
                color = QColor("#1c3d34")
                degisim_item.setForeground(QColor("#4ade80"))
            elif islem == "STOK ÇIKIŞI" or islem == "ÜRÜN SİLME":
                color = QColor("#4a1f2c")
                degisim_item.setForeground(QColor("#f87171"))
            else:
                color = QColor("#1e293b")
            for col in [4, 5]:
                item = self.rapor_tablosu.item(satir, col)
                if item: item.setBackground(color)
        self.rapor_tablosu.resizeRowsToContents()
        self.rapor_tablosu.setSortingEnabled(True)

    def menu_goster(self, pos):
        """Tablo satırına sağ tıklandığında 'Geri Al' menüsünü açar."""
        item = self.rapor_tablosu.itemAt(pos)
        if not item: return
        menu = QMenu(self)
        geri_al_action = QAction(" Bu İşlemi Geri Al ", self)
        row = item.row()
        tarih_item = self.rapor_tablosu.item(row, 0)
        hareket_id = tarih_item.data(Qt.ItemDataRole.UserRole)
        geri_al_action.triggered.connect(lambda: self.hareketi_geri_al(hareket_id))
        menu.addAction(geri_al_action)
        menu.exec(self.rapor_tablosu.viewport().mapToGlobal(pos))

    def hareketi_geri_al(self, hareket_id):
        """Seçilen işlemi geri almak için onay ister ve veritabanı metodunu çağırır."""
        msg = QMessageBox(self)
        msg.setWindowTitle("Geri Alma Onayı")
        msg.setText("Bu işlemi geri almak istediğinize emin misiniz?\nStok miktarı işlem öncesine döndürülecek.")
        msg.setIcon(QMessageBox.Icon.Question)

        # Butonları Türkçe ekliyoruz
        btn_evet = msg.addButton("Evet", QMessageBox.ButtonRole.YesRole)
        btn_hayir = msg.addButton("Hayır", QMessageBox.ButtonRole.NoRole)

        msg.setDefaultButton(btn_hayir)  # Yanlışlıkla basılmasın diye varsayılan 'Hayır'
        msg.exec()

        if msg.clickedButton() == btn_evet:
            basari, mesaj = self.veritabani.islem_geri_al(hareket_id, self.kullanici_adi)
            if basari:
                self.raporu_guncelle()
                QMessageBox.information(self, "Başarılı", mesaj)
            else:
                QMessageBox.warning(self, "Hata", mesaj)


class AnaStokSayfasi(QWidget):
    """
    Uygulamanın ana ekranıdır. Tüm ürünlerin listelendiği, arama/filtreleme
    yapılabildiği ve temel stok işlemlerinin (Ekle/Sil/Düzenle) yönetildiği merkezdir.
    """

    def __init__(self, veritabani_yoneticisi, status_bar, kullanici_adi):
        super().__init__()
        self.veritabani = veritabani_yoneticisi
        self.status_bar = status_bar
        self.kullanici_adi = kullanici_adi
        self.guncel_filtreler = {}
        self.float_validator = QDoubleValidator()
        self.float_validator.setBottom(0.0)
        self.float_validator.setNotation(QDoubleValidator.Notation.StandardNotation)
        self.arayuz_olustur()
        self.stogu_guncelle_arayuz()

    def create_metric_card(self, parent_layout, title, value, unit):
        """Ana sayfa üstündeki özet bilgi kutucuklarını oluşturur."""
        card_frame = QFrame()
        card_frame.setObjectName("metricCard")
        card_layout = QVBoxLayout(card_frame)
        card_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title_label = QLabel(title)
        title_label.setObjectName("metricTitle")
        if title == "Toplam Değer": title_label.setStyleSheet("padding-left: 16px;")
        value_layout = QHBoxLayout()
        value_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label = QLabel(value)
        value_label.setObjectName("metricValue")
        unit_label = QLabel(unit)
        unit_label.setObjectName("metricUnit")
        value_layout.addWidget(value_label)
        value_layout.addWidget(unit_label)
        # card_layout.addWidget(icon_label) # Icon removed
        card_layout.addWidget(title_label)
        card_layout.addLayout(value_layout)
        if not hasattr(self, 'metric_cards'): self.metric_cards = {}
        if title == "Ürün Çeşidi":
            self.metric_cards['urun_cesidi'] = value_label
        elif title == "Toplam Değer":
            self.metric_cards['toplam_deger'] = value_label
        elif title == "Düşük Stok":
            self.metric_cards['dusuk_stok'] = value_label
        parent_layout.addWidget(card_frame, 1)

    def arayuz_olustur(self):
        ana_duzen = QVBoxLayout(self)
        ana_duzen.setContentsMargins(25, 25, 25, 25)
        ana_duzen.setSpacing(20)
        dashboard_layout = QHBoxLayout()
        dashboard_layout.setSpacing(20)
        self.create_metric_card(dashboard_layout, "Ürün Çeşidi", "0", "adet")
        self.create_metric_card(dashboard_layout, "Toplam Değer", "0", "")
        self.create_metric_card(dashboard_layout, "Düşük Stok", "0", "ürün")
        ana_duzen.addLayout(dashboard_layout)
        ust_duzen = QHBoxLayout()
        ust_duzen.setSpacing(12)
        search_container = QFrame()
        search_container.setObjectName("searchContainer")
        search_layout = QHBoxLayout(search_container)
        search_layout.setContentsMargins(0, 0, 0, 0)
        search_layout.setSpacing(0)
        self.arama_input = QLineEdit()
        self.arama_input.setPlaceholderText("Ürün kodu, adı, kategori veya fiyat ile ara...")
        self.arama_input.setObjectName("searchInput")
        self.arama_input.setClearButtonEnabled(True)
        self.arama_input.textChanged.connect(lambda: self.tabloyu_filtrele())
        search_layout.addWidget(self.arama_input)
        ust_duzen.addWidget(search_container, 1)
        action_container = QHBoxLayout()
        self.filter_btn = QPushButton("Filtrele")
        self.filter_btn.setObjectName("filterBtn")
        self.filter_btn.clicked.connect(self.show_filter_dialog)
        self.yeni_urun_goster_btn = QPushButton("Yeni Ürün / Stok Ekle")
        self.yeni_urun_goster_btn.setObjectName("yeniUrunBtn")
        self.yeni_urun_goster_btn.clicked.connect(self.yeni_urun_formu_goster_gizle)
        action_container.addWidget(self.filter_btn)
        action_container.addWidget(self.yeni_urun_goster_btn)
        ust_duzen.addLayout(action_container)
        ana_duzen.addLayout(ust_duzen)
        self.ekleme_formu_frame = QFrame()
        self.ekleme_formu_frame.setObjectName("eklemeFormu")
        ekleme_duzen = QFormLayout(self.ekleme_formu_frame)
        self.yeni_urun_kodu_input = QLineEdit()
        self.yeni_urun_kodu_input.setMaxLength(13)
        self.yeni_urun_kodu_input.setValidator(QRegularExpressionValidator(QRegularExpression("[0-9]*")))
        self.yeni_urun_kodu_input.setPlaceholderText("13 Haneli Kod")
        self.yeni_urun_input = QLineEdit()
        self.yeni_kategori_input = QLineEdit()
        self.yeni_fiyat_input = QLineEdit()
        self.yeni_fiyat_input.setValidator(self.float_validator)
        miktar_layout = QHBoxLayout()
        self.yeni_miktar_input = QLineEdit()
        self.yeni_miktar_input.setValidator(self.float_validator)
        self.yeni_birim_input = QComboBox()
        self.yeni_birim_input.addItems(STANDART_BIRIMLER)
        self.yeni_birim_input.setFixedWidth(100)
        miktar_layout.addWidget(self.yeni_miktar_input, 1)
        miktar_layout.addWidget(self.yeni_birim_input)
        self.yeni_min_stok_input = QLineEdit("10")
        self.yeni_min_stok_input.setValidator(self.float_validator)
        self.yeni_skt_input = QDateEdit()
        self.yeni_skt_input.setCalendarPopup(True)
        self.yeni_skt_input.setDate(QDate.currentDate().addYears(1))
        self.yeni_skt_input.setDisplayFormat("yyyy-MM-dd")
        self.onayla_ekle_btn = QPushButton("Onayla")
        ekleme_duzen.addRow("Ürün Kodu (13 Hane):", self.yeni_urun_kodu_input)
        ekleme_duzen.addRow("Ürün Adı:", self.yeni_urun_input)
        ekleme_duzen.addRow("Kategori:", self.yeni_kategori_input)
        ekleme_duzen.addRow("Alış Fiyatı (Maliyet ₺):", self.yeni_fiyat_input)
        ekleme_duzen.addRow("Miktar / Birim:", miktar_layout)
        ekleme_duzen.addRow("Min. Stok:", self.yeni_min_stok_input)
        ekleme_duzen.addRow("Son Kul. Tarihi:", self.yeni_skt_input)
        ekleme_duzen.addRow(self.onayla_ekle_btn)
        form_notu = QLabel("Not: Mevcut bir 'Ürün Kodu' girerseniz, girilen miktar o ürünün stoğuna eklenecektir.")
        form_notu.setObjectName("pageSubtitle")
        ekleme_duzen.addRow(form_notu)
        self.ekleme_formu_frame.hide()
        ana_duzen.addWidget(self.ekleme_formu_frame)
        self.onayla_ekle_btn.clicked.connect(self.yeni_urun_ekle)
        # Tablo
        self.stok_tablosu = QTableWidget()
        self.stok_tablosu.setObjectName("stokTablosu")
        self.stok_tablosu.setColumnCount(10)
        self.stok_tablosu.setHorizontalHeaderLabels(
            ["Doluluk", "Ürün Kodu", "Ürün Adı", "Kategori", "Alış Fiyatı", "Miktar", "Birim", "Min. Stok", "SKT",
             "İşlemler"])
        header = self.stok_tablosu.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.stok_tablosu.setColumnWidth(0, 100)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        self.stok_tablosu.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stok_tablosu.setEditTriggers(QAbstractItemView.EditTrigger.DoubleClicked)
        self.stok_tablosu.itemChanged.connect(self.hucre_degisikligini_kaydet)
        self.stok_tablosu.setSortingEnabled(True)
        self.stok_tablosu.sortByColumn(2, Qt.SortOrder.AscendingOrder)
        self.stok_tablosu.verticalHeader().setVisible(False)
        ana_duzen.addWidget(self.stok_tablosu, 1)

    def guncelle_dashboard(self):
        veri = self.veritabani.genel_bakis_getir()
        if hasattr(self, 'metric_cards'):
            self.metric_cards['urun_cesidi'].setText(str(veri['urun_cesidi']))
            self.metric_cards['toplam_deger'].setText(f"{veri['toplam_deger']:,.2f}₺")
            self.metric_cards['dusuk_stok'].setText(str(veri['dusuk_stok']))
            self.metric_cards['dusuk_stok'].setProperty("lowStock", veri['dusuk_stok'] > 0)
            self.metric_cards['dusuk_stok'].style().polish(self.metric_cards['dusuk_stok'])

    def stogu_guncelle_arayuz(self, urun_listesi=None):
        """
        Tablodaki verileri veritabanından çekerek yeniler.
        Ayrıca dashboard kartlarındaki özet verileri de günceller.
        """
        self.guncelle_dashboard()
        try:
            self.stok_tablosu.itemChanged.disconnect(self.hucre_degisikligini_kaydet)
        except TypeError:
            pass
        self.stok_tablosu.setSortingEnabled(False)
        self.stok_tablosu.setRowCount(0)
        if urun_listesi is None: urun_listesi = self.veritabani.urunleri_getir()
        bugun = date.today()
        self.stok_tablosu.setRowCount(len(urun_listesi))
        self.stok_tablosu.setRowCount(len(urun_listesi))
        for satir, (id_val, urun_kodu, ad, kategori, fiyat, miktar, birim, min_stok, baslangic_miktari,
                    son_kullanma_tarihi) in enumerate(urun_listesi):
            progress_bar_widget = QWidget()
            progress_layout = QHBoxLayout(progress_bar_widget)
            progress_layout.setContentsMargins(5, 2, 5, 2)
            progress_bar = QProgressBar()
            progress_bar.setRange(0, 100)
            progress_bar.setTextVisible(False)
            progress_bar.setFixedHeight(12)
            if baslangic_miktari > 0:
                oran = int((miktar / baslangic_miktari) * 100)
                oran = min(100, max(0, oran))
            else:
                oran = 0
            progress_bar.setValue(oran)
            stili = """
                QProgressBar { border: 1px solid #334155; border-radius: 4px; background-color: #1e293b; }
                QProgressBar::chunk { border-radius: 4px; }
            """
            if oran < 20:
                stili += "QProgressBar::chunk { background-color: #ef4444; }"
            elif oran < 50:
                stili += "QProgressBar::chunk { background-color: #eab308; }"
            else:
                stili += "QProgressBar::chunk { background-color: #22c55e; }"
            progress_bar.setStyleSheet(stili)
            progress_layout.addWidget(progress_bar)
            kod_item = QTableWidgetItem(urun_kodu)
            kod_item.setData(Qt.ItemDataRole.UserRole, id_val)
            ad_item = QTableWidgetItem(ad)
            kategori_item = QTableWidgetItem(kategori)
            fiyat_item = NumericTableWidgetItem(f"{fiyat:.2f}₺", fiyat)
            miktar_str = f"{miktar:g}"
            min_stok_str = f"{min_stok:g}"
            miktar_item = NumericTableWidgetItem(miktar_str, miktar)
            birim_item = QTableWidgetItem(birim)
            min_stok_item = NumericTableWidgetItem(min_stok_str, min_stok)
            skt_item = QTableWidgetItem(son_kullanma_tarihi if son_kullanma_tarihi else "-")
            kod_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            fiyat_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            miktar_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            birim_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            min_stok_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            skt_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            kod_item.setFlags(kod_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            ad_item.setFlags(ad_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            miktar_item.setFlags(miktar_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            birim_item.setFlags(birim_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            min_stok_item.setFlags(min_stok_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            skt_item.setFlags(skt_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self.stok_tablosu.setCellWidget(satir, 0, progress_bar_widget)
            self.stok_tablosu.setItem(satir, 1, kod_item)
            self.stok_tablosu.setItem(satir, 2, ad_item)
            self.stok_tablosu.setItem(satir, 3, kategori_item)
            self.stok_tablosu.setItem(satir, 4, fiyat_item)
            self.stok_tablosu.setItem(satir, 5, miktar_item)
            self.stok_tablosu.setItem(satir, 6, birim_item)
            self.stok_tablosu.setItem(satir, 7, min_stok_item)
            self.stok_tablosu.setItem(satir, 8, skt_item)
            bg_color = None
            if son_kullanma_tarihi:
                try:
                    skt_date = datetime.strptime(son_kullanma_tarihi, "%Y-%m-%d").date()
                    kalan_gun = (skt_date - bugun).days
                    if kalan_gun < 0:
                        bg_color = QColor("#7f1d1d")
                        skt_item.setText(f"GECİKMİŞ ({son_kullanma_tarihi})")
                        skt_item.setForeground(QColor("#fca5a5"))
                    elif kalan_gun <= 30:
                        skt_item.setForeground(QColor("#fbbf24"))
                except ValueError:
                    pass
            if bg_color is None and miktar <= min_stok:
                bg_color = QColor("#4a1f2c")
                miktar_item.setText(f"{miktar_str}")
                miktar_item.setForeground(QColor("#f87171"))
            if bg_color:
                for col in range(1, self.stok_tablosu.columnCount() - 1):
                    item = self.stok_tablosu.item(satir, col)
                    if item: item.setBackground(bg_color)
            menu_btn = QPushButton("•••")
            menu_btn.setObjectName("menuButton")
            menu_btn.setCursor(Qt.CursorShape.PointingHandCursor)
            menu_btn.clicked.connect(lambda checked, u_id=id_val, u_ad=ad: self.guncelle_menusu_goster(u_id, u_ad))
            btn_container = QWidget()
            btn_layout = QHBoxLayout(btn_container)
            btn_layout.setContentsMargins(0, 0, 0, 0)
            btn_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
            btn_layout.addWidget(menu_btn)
            btn_layout.addWidget(menu_btn)
            self.stok_tablosu.setCellWidget(satir, 9, btn_container)
            self.stok_tablosu.setRowHeight(satir, 40)
            self.stok_tablosu.setRowHeight(satir, 50)
        self.stok_tablosu.resizeRowsToContents()
        self.stok_tablosu.setSortingEnabled(True)
        self.tabloyu_filtrele()
        self.stok_tablosu.itemChanged.connect(self.hucre_degisikligini_kaydet)

    def guncelle_menusu_goster(self, urun_id, urun_adi_gosterim):
        menu = QMenu(self)
        duzenle = QAction("Bilgileri Düzenle (Detay)", self)
        duzenle.triggered.connect(lambda: self.urun_duzenle(urun_id))
        artir = QAction("Stok Artır", self)
        artir.triggered.connect(lambda: self.miktar_girdi_goster(urun_id, urun_adi_gosterim, 'artır'))
        satis = QAction("Satış Yap", self)
        satis.triggered.connect(lambda: self.miktar_girdi_goster(urun_id, urun_adi_gosterim, 'satis'))
        sil = QAction("Ürünü Sil", self)
        sil.setObjectName("dangerAction")
        sil.triggered.connect(lambda: self.urun_sil(urun_id, urun_adi_gosterim))
        menu.addAction(duzenle)
        menu.addSeparator()
        menu.addAction(artir)
        menu.addAction(satis)
        menu.addSeparator()
        menu.addAction(sil)
        menu.exec(QCursor.pos())

    def urun_duzenle(self, urun_id):
        urun_detaylari = self.veritabani.urun_detay_getir(urun_id)
        if not urun_detaylari: return
        dialog = UrunDuzenlemeDialog(urun_detaylari, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            veri = dialog.get_data()
            if veri:
                id_val, ad, kat, fiyat, birim, min_stok, yeni_skt = veri
                basari, mesaj = self.veritabani.urun_detay_guncelle(id_val, ad, kat, fiyat, birim, min_stok, yeni_skt,
                                                                    self.kullanici_adi)
                if basari:
                    self.stogu_guncelle_arayuz()
                    self.status_bar.showMessage(f"'{ad}' güncellendi.", 3000)
                else:
                    QMessageBox.warning(self, "Güncelleme Hatası", mesaj)

    def yeni_urun_ekle(self):
        kod = self.yeni_urun_kodu_input.text()
        if len(kod) != 13:
            QMessageBox.warning(self, "Uyarı", "Ürün Kodu 13 haneli olmalıdır.")
            return
        ad = self.yeni_urun_input.text().strip().upper()
        kat = self.yeni_kategori_input.text().strip().upper()
        f_str = self.yeni_fiyat_input.text().strip()
        m_str = self.yeni_miktar_input.text().strip()
        min_s_str = self.yeni_min_stok_input.text().strip()
        birim = self.yeni_birim_input.currentText()
        skt_tarihi = self.yeni_skt_input.date().toString("yyyy-MM-dd")
        if not m_str:
            QMessageBox.warning(self, "Uyarı", "Miktar alanı zorunludur.")
            return
        mevcut_urun = self.veritabani.cursor.execute("SELECT id FROM urunler WHERE urun_kodu = ?", (kod,)).fetchone()
        if not mevcut_urun and not all([ad, f_str, min_s_str]):
            QMessageBox.warning(self, "Uyarı",
                                "Yeni ürün eklerken (kod mevcut değilse) Ad, Alış Fiyatı ve Min. Stok alanları zorunludur.")
            return
        try:
            m = float(m_str.replace(',', '.'))
            f = float(f_str.replace(',', '.')) if f_str else 0.0
            min_s = float(min_s_str.replace(',', '.')) if min_s_str else 10.0
            if not (m > 0 and f >= 0 and min_s >= 0): raise ValueError()
        except(ValueError):
            QMessageBox.warning(self, "Hata", "Miktar, Fiyat ve Min. Stok pozitif sayı olmalıdır.")
            return
        basari, mesaj = self.veritabani.urun_ekle(kod, ad, kat, f, m, birim, min_s, skt_tarihi, self.kullanici_adi)
        if basari:
            self.stogu_guncelle_arayuz()
            self.yeni_urun_formu_goster_gizle()
            self.status_bar.showMessage(mesaj, 3000)
            for item in [self.yeni_urun_kodu_input, self.yeni_urun_input, self.yeni_kategori_input,
                         self.yeni_fiyat_input, self.yeni_miktar_input]:
                item.clear()
            self.yeni_min_stok_input.setText("10")
            self.yeni_birim_input.setCurrentIndex(0)
            self.yeni_skt_input.setDate(QDate.currentDate().addYears(1))
        else:
            QMessageBox.warning(self, "Ekleme Hatası", mesaj)

    def show_filter_dialog(self):
        dialog = FiltreDialog(self.guncel_filtreler, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            if dialog.reset_requested:
                self.guncel_filtreler = {}
                self.status_bar.showMessage("Filtreler temizlendi.", 3000)
                self.filter_btn.setText("Filtrele")
                self.filter_btn.setProperty("filtered", False)
            else:
                self.guncel_filtreler = dialog.get_filtreler()
                self.status_bar.showMessage("Gelişmiş filtreler uygulandı.", 3000)
                self.filter_btn.setText("Filtreleniyor")
                self.filter_btn.setProperty("filtered", True)
            self.filter_btn.style().polish(self.filter_btn)
        self.tabloyu_filtrele()

    def tabloyu_filtrele(self):
        metin_lower = self.arama_input.text().lower()
        min_f, max_f = self.guncel_filtreler.get("min_fiyat"), self.guncel_filtreler.get("max_fiyat")
        min_s, max_s = self.guncel_filtreler.get("min_stok"), self.guncel_filtreler.get("max_stok")
        dusuk_stok_only = self.guncel_filtreler.get("dusuk_stok_only", False)
        for i in range(self.stok_tablosu.rowCount()):
            urun_kodu = self.stok_tablosu.item(i, 1).text().lower()
            urun_adi = self.stok_tablosu.item(i, 2).text().lower()
            kategori = self.stok_tablosu.item(i, 3).text().lower()
            fiyat_text = self.stok_tablosu.item(i, 4).text().lower()
            text_match = (
                    metin_lower in urun_kodu or metin_lower in urun_adi or metin_lower in kategori or metin_lower in fiyat_text)
            fiyat = self.stok_tablosu.item(i, 4).sort_key
            miktar = self.stok_tablosu.item(i, 5).sort_key
            min_stok_val = self.stok_tablosu.item(i, 7).sort_key
            fiyat_match = (min_f is None or fiyat >= min_f) and (max_f is None or fiyat <= max_f)
            stok_match = (min_s is None or miktar >= min_s) and (max_s is None or miktar <= max_s)
            dusuk_stok_match = not dusuk_stok_only or miktar <= min_stok_val
            is_visible = text_match and fiyat_match and stok_match and dusuk_stok_match
            self.stok_tablosu.setRowHidden(i, not is_visible)

    def yeni_urun_formu_goster_gizle(self):
        if self.ekleme_formu_frame.isVisible():
            self.ekleme_formu_frame.hide()
        else:
            self.ekleme_formu_frame.show()
            self.yeni_urun_kodu_input.setFocus()

    def miktar_girdi_goster(self, urun_id, urun_adi, mod):
        detaylar = self.veritabani.urun_detay_getir(urun_id)
        if not detaylar: return
        mevcut = detaylar[5]
        birim = detaylar[6]
        guncel_alis_fiyati = detaylar[4]

        if mod == 'artır':
            fiil = "Artır"

            #  Özel Buton İsimli Dialog
            dialog = QInputDialog(self)
            dialog.setWindowTitle(f"Stok {fiil} - [{urun_adi}]")
            dialog.setLabelText(f"Mevcut Miktar: {mevcut:g} {birim}\nLütfen eklenecek miktarı girin:")

            # Sayısal giriş ayarları
            dialog.setInputMode(QInputDialog.InputMode.DoubleInput)
            dialog.setDoubleRange(0.0, 999999.0)
            dialog.setDoubleDecimals(2)
            dialog.setDoubleValue(1.0)

            # Buton metinlerini değiştirme
            dialog.setOkButtonText("Tamam")
            dialog.setCancelButtonText("İptal Et")

            # Pencereyi aç ve sonucu kontrol et
            if dialog.exec() == QDialog.DialogCode.Accepted:
                m = dialog.doubleValue()
                self.stok_miktari_degistir(urun_id, urun_adi, m)


        elif mod == 'satis':
            dialog = SatisDialog(urun_adi, mevcut, birim, guncel_alis_fiyati, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                m, satis_fiyati = dialog.get_values()
                if m is not None: self.stok_miktari_degistir(urun_id, urun_adi, -m, satis_fiyati)

    def stok_miktari_degistir(self, urun_id, urun_adi, m_farki, satis_fiyati=None):
        basari, mesaj = self.veritabani.urun_miktar_guncelle(urun_id, m_farki, self.kullanici_adi, satis_fiyati)
        if basari:
            self.stogu_guncelle_arayuz()
            self.status_bar.showMessage(f"'{urun_adi}' stoğu güncellendi.", 3000)
        else:
            QMessageBox.warning(self, "Uyarı", mesaj)

    def urun_sil(self, urun_id, urun_adi):
        # Standart soru kutusu yerine özel kutu oluşturuyoruz
        msg = QMessageBox(self)
        msg.setWindowTitle("Silmeyi Onayla")
        msg.setText(f"<b>{urun_adi}</b> (ID: {urun_id}) ürününü silmek istediğinizden emin misiniz?")
        msg.setIcon(QMessageBox.Icon.Question)

        # Butonları Türkçe metinlerle ekliyoruz
        btn_evet = msg.addButton("Evet", QMessageBox.ButtonRole.YesRole)
        btn_hayir = msg.addButton("Hayır", QMessageBox.ButtonRole.NoRole)

        # Varsayılan olarak 'Hayır' seçili gelsin (Güvenlik için)
        msg.setDefaultButton(btn_hayir)

        msg.exec()

        # Eğer basılan buton 'Evet' ise silme işlemini yap
        if msg.clickedButton() == btn_evet:
            self.veritabani.urun_sil(urun_id, self.kullanici_adi)
            self.stogu_guncelle_arayuz()
            self.status_bar.showMessage(f"'{urun_adi}' silindi.", 3000)

    def verileri_disa_aktar(self):
        """
        Mevcut stok listesini Excel (.xlsx) veya CSV formatında dışa aktarır.
        Excel için 'openpyxl' kütüphanesini kullanır.
        """
        # Kullanıcıya Excel mi CSV mi seçtirmek için filtreleri ayarlıyoruz
        filtreler = "Excel Dosyası (*.xlsx);;CSV Dosyası (*.csv)"
        kayit_yolu, secilen_filtre = QFileDialog.getSaveFileName(
            self,
            "Verileri Dışa Aktar",
            "stok_raporu",  # Varsayılan dosya adı (uzantısız)
            filtreler
        )

        if not kayit_yolu:
            return

        try:
            urunler = self.veritabani.urunleri_getir()
            basliklar = ["ID", "Urun Kodu", "Ad", "Kategori", "Fiyat", "Miktar", "Birim", "Min. Stok",
                         "Baslangic Miktari", "Son Kul. Tarihi"]

            # --- EXCEL SEÇİLDİYSE ---
            if ".xlsx" in kayit_yolu:
                if not EXCEL_MUMKUN:
                    QMessageBox.warning(self, "Eksik Kütüphane",
                                        "Excel oluşturmak için 'openpyxl' kütüphanesi yüklü olmalıdır.\nLütfen 'pip install openpyxl' komutunu çalıştırın.")
                    return

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Stok Listesi"

                # 1. Başlıkları Yaz ve Stille
                ws.append(basliklar)

                # Başlık satırını (1. satır) kalın yap ve arka plan rengi ver
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="3b82f6", end_color="3b82f6", fill_type="solid")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # 2. Verileri Yaz
                for row in urunler:
                    ws.append(row)

                # 3. Sütun Genişliklerini Otomatik Ayarla (İsteğe Bağlı Estetik)
                for column_cells in ws.columns:
                    length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

                wb.save(kayit_yolu)
                self.status_bar.showMessage(f"Veriler Excel olarak '{kayit_yolu}' dosyasına aktarıldı.", 5000)

            # --- CSV SEÇİLDİYSE ---
            else:
                with open(kayit_yolu, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f, delimiter=';')  # Excel'in CSV'yi düzgün açması için ; kullanılabilir
                    writer.writerow(basliklar)
                    writer.writerows(urunler)
                self.status_bar.showMessage(f"Veriler CSV olarak '{kayit_yolu}' dosyasına aktarıldı.", 5000)

        except Exception as e:
            QMessageBox.critical(self, "Dışa Aktarma Hatası", f"Dosya yazılırken bir hata oluştu:\n{e}")

    def hucre_degisikligini_kaydet(self, item):
        """
        Tablo üzerinde hücreye çift tıklayarak yapılan değişiklikleri algılar ve kaydeder.
        Sadece Fiyat ve Kategori sütunları düzenlenebilir.
        """
        satir, sutun = item.row(), item.column()
        try:
            urun_kodu_item = self.stok_tablosu.item(satir, 1)
            urun_id = urun_kodu_item.data(Qt.ItemDataRole.UserRole)
        except Exception:
            return
        yeni_deger_str = item.text()
        if sutun == 3:
            basari, mesaj = self.veritabani.urun_hucre_guncelle(urun_id, "kategori", yeni_deger_str.upper(),
                                                                self.kullanici_adi)
        elif sutun == 4:
            try:
                yeni_fiyat = float(yeni_deger_str.replace('₺', '').replace(',', '.').strip())
                if yeni_fiyat < 0: raise ValueError("Fiyat negatif olamaz")
                basari, mesaj = self.veritabani.urun_hucre_guncelle(urun_id, "fiyat", yeni_fiyat, self.kullanici_adi)
                if basari:
                    self.guncelle_dashboard()
                    self.stok_tablosu.itemChanged.disconnect(self.hucre_degisikligini_kaydet)
                    self.stok_tablosu.setItem(satir, sutun, NumericTableWidgetItem(f"{yeni_fiyat:.2f}₺", yeni_fiyat))
                    self.stok_tablosu.itemChanged.connect(self.hucre_degisikligini_kaydet)
            except ValueError:
                basari, mesaj = False, "Lütfen geçerli bir pozitif sayı girin."
        else:
            return
        if not basari:
            QMessageBox.warning(self, "Hata", mesaj)
            self.stogu_guncelle_arayuz()


# =============================================================================
# 5. DASHBOARD AND CUSTOM WIDGETS
# =============================================================================

class ThemeToggleButton(QWidget):
    """
    Aydınlık/Karanlık mod arasında geçiş sağlayan animasyonlu anahtar (Toggle Switch).
    Güneş ve Ay ikonları ile görselleştirilmiştir.
    """
    toggled = pyqtSignal(bool)  # Signal to emit when toggled (True=Dark, False=Light)

    def __init__(self, parent=None, is_dark=True):
        super().__init__(parent)
        self.setFixedSize(60, 30)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        self._is_dark = is_dark

        # Animasyon için değişken (0.0 = Sol/Gündüz, 1.0 = Sağ/Gece)
        self._position = 1.0 if is_dark else 0.0

        # Animasyon Tanımı
        self.animation = QPropertyAnimation(self, b"position")
        self.animation.setDuration(300)
        self.animation.setEasingCurve(QEasingCurve.Type.InOutCubic)  # Smooth geçiş

    @pyqtProperty(float)
    def position(self):
        return self._position

    @position.setter
    def position(self, pos):
        self._position = pos
        self.update()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.toggle()
        super().mousePressEvent(event)

    def toggle(self):
        # Durumu tersine çevir
        self._is_dark = not self._is_dark

        # Animasyonu başlat
        start_val = self._position
        end_val = 1.0 if self._is_dark else 0.0

        self.animation.setStartValue(start_val)
        self.animation.setEndValue(end_val)
        self.animation.start()

        self.toggled.emit(self._is_dark)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # Arka Plan Renkleri (Gök Mavisi <-> Koyu Mavi)
        # position 0.0 = Gündüz, 1.0 = Gece
        if self._position > 0.5:
            bg_color = QColor("#0f172a")  # Gece arkaplanı
            border_color = QColor("#334155")
        else:
            bg_color = QColor("#bfdbfe")  # Gündüz arkaplanı (Açık Mavi)
            border_color = QColor("#60a5fa")

        # 1. Kapsül Arka Plan
        rect = self.rect()
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QBrush(bg_color))
        painter.drawRoundedRect(rect, 15, 15)

        # İç çerçeve (opsiyonel detay)
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.setPen(QPen(border_color, 1))
        painter.drawRoundedRect(rect.adjusted(1, 1, -1, -1), 14, 14)

        # 2. Hareketli Yuvarlak (Thumb)
        # Margin: 4px
        margin = 4
        thumb_size = self.height() - (margin * 2)  # 22px

        # X Konumu: 0.0 iken solda, 1.0 iken sağda
        travel_dist = self.width() - thumb_size - (margin * 2)
        thumb_x = margin + (self._position * travel_dist)
        thumb_y = margin

        container_rect = QRectF(thumb_x, thumb_y, thumb_size, thumb_size)

        # Thumb Kendisi
        painter.setPen(Qt.PenStyle.NoPen)
        painter.setBrush(QColor("#ffffff"))
        painter.drawEllipse(container_rect)

        # 3. İkon (Thumb'ın üzerine çiz)
        # İkon her zaman thumb'ın merkezinde olacak
        # Hangi ikon? Gündüzse (Pos < 0.5) Güneş, Geceyse (Pos > 0.5) Ay

        painter.save()
        # İkonu merkeze hizalamak için translate et
        center = container_rect.center()
        painter.translate(center)

        if self._position < 0.5:
            # --- GÜNEŞ İKONU ---
            # Merkez Daire
            painter.setBrush(QColor("#f59e0b"))  # Turuncu
            painter.drawEllipse(QPoint(0, 0), 5, 5)

            # Işınlar
            painter.setPen(QPen(QColor("#f59e0b"), 2, Qt.PenStyle.SolidLine, Qt.PenCapStyle.RoundCap))
            for i in range(8):
                angle = 45 * i
                painter.rotate(angle)
                painter.drawLine(0, 7, 0, 9)
                painter.rotate(-angle)

        else:
            # --- AY İKONU ---
            # Hilal
            painter.setBrush(QColor("#6366f1"))  # Morumsu Mavi
            # İlk daire
            painter.drawEllipse(QPoint(-1, -1), 6, 6)

            # Kesik (Maske) - Arka plan rengiyle aynı veya beyaz yaparak 'hilal' görünümü veremeyiz çünkü thumb beyaz.
            # Thumb beyaz olduğu için maske 'beyaz' olmalı.
            painter.setBrush(QColor("#ffffff"))
            painter.drawEllipse(QPoint(2, -2), 5, 5)

        painter.restore()


class SimpleChartWidget(QWidget):
    """
    Basit grafikler çizen özel widget.
    Temaya uyumlu olması için renkleri dinamik alır.
    """
    clicked = pyqtSignal()

    def __init__(self, chart_type="line", data=None, title="", is_dark=True, parent=None):
        super().__init__(parent)
        self.chart_type = chart_type
        self.data = data if data else {}
        self.title = title
        self.setMinimumHeight(250)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

        # Başlangıç tema renklerini ayarla
        self.set_theme(is_dark)

    def set_theme(self, is_dark):
        """Temaya göre grafik renklerini günceller."""
        self.is_dark = is_dark
        if is_dark:
            self.text_color = QColor("#94a3b8")
            self.axis_color = QColor("#334155")
            self.grid_color = QColor("#1e293b")
        else:
            self.text_color = QColor("#475569")  # Koyu gri (Açık tema için)
            self.axis_color = QColor("#cbd5e1")  # Açık gri sınır
            self.grid_color = QColor("#f1f5f9")
        self.update()  # Yeniden çiz

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

    def paintEvent(self, event):
        try:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)

            # Başlık Rengi (Dinamik)
            painter.setPen(self.text_color)
            painter.setFont(QFont("Segoe UI", 10, QFont.Weight.Bold))
            painter.drawText(self.rect().adjusted(10, 10, -10, -10),
                             Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft, self.title)

            if not self.data:
                painter.drawText(self.rect(), Qt.AlignmentFlag.AlignCenter, "Veri Yok")
                return

            margin_left = 40
            margin_bottom = 50
            margin_top = 40
            margin_right = 20
            chart_rect = self.rect().adjusted(margin_left, margin_top, -margin_right, -margin_bottom)

            if self.chart_type == "line":
                self.draw_line_chart(painter, chart_rect)
            elif self.chart_type == "bar":
                self.draw_bar_chart(painter, chart_rect)
            elif self.chart_type == "pie":
                self.draw_pie_chart(painter, chart_rect)
        except Exception as e:
            print(f"Paint Error: {e}")

    # --- ALT METOTLARDA RENKLERİ GÜNCELLEYELİM ---
    def draw_line_chart(self, painter, rect):
        # Eksenler (Dinamik Renk)
        painter.setPen(QPen(self.axis_color, 2))
        painter.drawLine(int(rect.left()), int(rect.bottom()), int(rect.right()), int(rect.bottom()))
        painter.drawLine(int(rect.left()), int(rect.top()), int(rect.left()), int(rect.bottom()))

        values = list(self.data.values())
        keys = list(self.data.keys())
        if not values: return
        max_val = max(values) if max(values) > 0 else 10
        min_val = 0
        points = []
        step_x = rect.width() / (len(values) - 1) if len(values) > 1 else rect.width()

        for i, val in enumerate(values):
            x = rect.left() + (i * step_x)
            ratio = (val - min_val) / (max_val - min_val)
            y = rect.bottom() - (ratio * rect.height())
            points.append(QPoint(int(x), int(y)))
            if len(values) > 10 and i % 3 != 0: continue

            # X Ekseni Yazıları (Dinamik Renk)
            painter.setPen(self.text_color)
            painter.setFont(QFont("Segoe UI", 8))
            painter.drawText(int(x) - 15, int(rect.bottom()) + 5, 30, 20, Qt.AlignmentFlag.AlignCenter, str(keys[i]))

        painter.setPen(QPen(QColor("#3b82f6"), 3))
        for i in range(len(points) - 1):
            painter.drawLine(points[i], points[i + 1])

    def draw_bar_chart(self, painter, rect):
        values = list(self.data.values())
        keys = list(self.data.keys())
        if not values: return
        max_val = max(values) if max(values) > 0 else 10
        count = len(values)
        bar_width = (rect.width() / count) * 0.6
        spacing = (rect.width() / count) * 0.4
        start_x = rect.left() + (spacing / 2)

        for i, val in enumerate(values):
            ratio = val / max_val
            bar_height = rect.height() * ratio
            x = start_x + (i * (bar_width + spacing))
            y = rect.bottom() - bar_height
            bar_rect = QRectF(x, y, bar_width, bar_height)

            painter.setBrush(QBrush(QColor("#10b981")))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawRoundedRect(bar_rect, 4, 4)

            # Bar Altı Yazıları (Dinamik Renk)
            painter.setFont(QFont("Segoe UI", 8))
            painter.setPen(self.text_color)
            painter.drawText(QRectF(x - 5, rect.bottom() + 5, bar_width + 10, 40),
                             Qt.AlignmentFlag.AlignCenter | Qt.TextFlag.TextWordWrap,
                             str(keys[i]))

    def draw_pie_chart(self, painter, rect):
        values = list(self.data.values())
        keys = list(self.data.keys())
        if not values: return

        total = sum(values)
        if total == 0: return

        # Lejant için alan kontrolü
        # Eğer genişlik 400'den büyükse sağa lejant ekle
        show_legend = rect.width() > 380

        # Gelişmiş Renk Paleti (Daha modern renkler)
        colors = [
            QColor("#3b82f6"), QColor("#10b981"), QColor("#f59e0b"),
            QColor("#ef4444"), QColor("#8b5cf6"), QColor("#ec4899"),
            QColor("#06b6d4"), QColor("#6366f1")
        ]

        start_angle = 90 * 16  # 12 yönünden başla

        if show_legend:
            # Lejant var: Pastayı sola kaydır
            legend_width = 140
            chart_width = rect.width() - legend_width
            center_x = rect.left() + (chart_width / 2)
            center_y = rect.center().y()
            radius = min(chart_width, rect.height()) / 2 * 0.8

            # Lejantı Çiz
            legend_x = rect.left() + chart_width + 10
            legend_y = rect.top() + 20

            painter.setFont(QFont("Segoe UI", 9))

            for i, (key, val) in enumerate(zip(keys, values)):
                color = colors[i % len(colors)]

                # Renk Kutusu
                box_rect = QRectF(legend_x, legend_y + (i * 25), 12, 12)
                painter.setBrush(QBrush(color))
                painter.setPen(Qt.PenStyle.NoPen)
                painter.drawRoundedRect(box_rect, 3, 3)

                # Yazı
                painter.setPen(self.text_color)
                text_rect = QRectF(legend_x + 18, legend_y + (i * 25) - 4, legend_width - 20, 20)
                # Yüzdeli gösterim
                perc = (val / total) * 100
                label_text = f"{key} (%{perc:.1f})"
                painter.drawText(text_rect, Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter, label_text)

        else:
            # Lejant yok: Ortala
            center_x = rect.center().x()
            center_y = rect.center().y()
            radius = min(rect.width(), rect.height()) / 2 * 0.8

        pie_rect = QRectF(center_x - radius, center_y - radius, radius * 2, radius * 2)

        for i, val in enumerate(values):
            span_angle = int((val / total) * 360 * 16)

            # Dilimi Çiz
            color = colors[i % len(colors)]
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.PenStyle.NoPen)
            painter.drawPie(pie_rect, start_angle, -span_angle)  # Saat yönünde gitmesi için negatif

            # Dilimin ortası için açı hesapla (Etiket konumu)
            mid_angle = start_angle + (-span_angle / 2)
            # Radyana çevir
            angle_rad = math.radians(mid_angle / 16)

            # Etiketin konumu (Merkezden biraz dışarıda)
            label_radius = radius * 0.65
            lbl_x = center_x + label_radius * math.cos(angle_rad)
            lbl_y = center_y - label_radius * math.sin(angle_rad)  # Y ekseni ters

            # Yüzdeyi hesapla
            percentage = (val / total) * 100
            if percentage > 5:  # Sadece %5'ten büyükse etiket koy
                text = f"%{percentage:.0f}"

                # Metni Yaz (Açık renk font çünkü dilimler genellikle koyu/canlı)
                painter.setPen(QColor("#ffffff"))
                painter.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))

                # Metni ortalayarak çiz
                text_rect = QRectF(lbl_x - 20, lbl_y - 10, 40, 20)
                painter.drawText(text_rect, Qt.AlignmentFlag.AlignCenter, text)

            start_angle -= span_angle
        # Şimdilik sadece dilimler ve yüzdeler yeterli.


class ChartDetailDialog(QDialog):
    """
    Dashboard grafiklerine tıklandığında açılan, grafiği daha büyük gösteren
    ve verileri tablo halinde sunan detay penceresi.
    """

    def __init__(self, chart_type, title, data, is_dark=True, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Detay: {title}")
        self.setMinimumSize(900, 600)

        # --- TEMA RENKLERİ ---
        if is_dark:
            bg_main = "#0f172a"
            text_main = "#e2e8f0"
            bg_card = "#1e293b"
            border_color = "#334155"
            header_text = "#f8fafc"
            table_header_bg = "#0f172a"
            table_header_text = "#94a3b8"
            btn_close_bg = "#ef4444"
        else:
            bg_main = "#f1f5f9"
            text_main = "#1e293b"
            bg_card = "#ffffff"
            border_color = "#cbd5e1"
            header_text = "#1e293b"
            table_header_bg = "#f8fafc"
            table_header_text = "#64748b"
            btn_close_bg = "#ef4444"

        self.setStyleSheet(f"background-color: {bg_main}; color: {text_main};")

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(20)

        # Başlık
        header = QLabel(title)
        header.setStyleSheet(f"font-size: 24px; font-weight: bold; color: {header_text};")
        layout.addWidget(header)

        # Grafik (Büyük)
        chart_frame = QFrame()
        chart_frame.setStyleSheet(
            f"background-color: {bg_card}; border-radius: 12px; border: 1px solid {border_color};")
        chart_layout = QVBoxLayout(chart_frame)
        self.chart = SimpleChartWidget(chart_type, data, "", is_dark=is_dark)  # Başlık zaten yukarıda, temayı aktar
        self.chart.setMinimumHeight(400)
        chart_layout.addWidget(self.chart)
        layout.addWidget(chart_frame, 2)

        # Veri Tablosu
        self.table = QTableWidget()
        self.table.setColumnCount(2)
        self.table.setHorizontalHeaderLabels(["Etiket", "Değer"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet(f"""
            QTableWidget {{ background-color: {bg_card}; border: 1px solid {border_color}; border-radius: 8px; }}
            QHeaderView::section {{ background-color: {table_header_bg}; color: {table_header_text}; border: none; padding: 8px; }}
            QTableWidget::item {{ padding: 8px; border-bottom: 1px solid {border_color}; color: {text_main}; }}
            QScrollBar:vertical {{ border: none; background: {bg_main}; width: 10px; margin: 0; }}
            QScrollBar::handle:vertical {{ background: {border_color}; min-height: 20px; border-radius: 5px; }}
            QScrollBar::handle:vertical:hover {{ background: #475569; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0px; }}
        """)

        # Verileri doldur
        row = 0
        for k, v in data.items():
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(str(k)))

            val_str = f"{v:,.2f}" if isinstance(v, float) else str(v)
            self.table.setItem(row, 1, QTableWidgetItem(val_str))
            row += 1

        layout.addWidget(self.table, 1)

        # Kapat Butonu
        btn_close = QPushButton("Kapat")
        btn_close.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_close.setStyleSheet(
            f"background-color: {btn_close_bg}; color: white; padding: 10px; border-radius: 6px; font-weight: bold;")
        btn_close.clicked.connect(self.close)
        layout.addWidget(btn_close, alignment=Qt.AlignmentFlag.AlignRight)


class ClickableCard(QFrame):
    """
    Tıklanabilir özellik kazandırılmış QFrame.
    Dashboard üzerindeki özet kartların sayfa yönlendirmesi yapabilmesi için kullanılır.
    """
    clicked = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setAttribute(Qt.WidgetAttribute.WA_Hover)

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton:
            self.clicked.emit()
        super().mousePressEvent(event)


class DashboardPage(QWidget):
    """
    Uygulamanın başlangıç ekranı (Kontrol Paneli).
    Kritik metrikleri, grafikleri ve özet bilgileri bir arada sunar.
    """
    navigation_requested = pyqtSignal(int)  # Sinyal: Ana pencereye gitmesi gereken sayfa indexini bildir

    def __init__(self, veritabani_yoneticisi, parent=None):
        super().__init__(parent)
        self.veritabani = veritabani_yoneticisi
        self.setup_ui()
        self.refresh_data()

    def setup_ui(self):
        # Ana layout'u oluştur (ScrollArea içerecek)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        # ScrollArea oluştur
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

        # Scroll'un içindeki Widget (Tüm içerik burada olacak)
        content_widget = QWidget()
        content_widget.setObjectName("dashboardContent")  # Styling için

        # İçerik layout'u (Eski layout buraya taşındı)
        layout = QVBoxLayout(content_widget)
        layout.setContentsMargins(30, 30, 30, 30)
        layout.setSpacing(25)

        # --- Üst Bar (Yenile Butonu) ---
        top_bar_layout = QHBoxLayout()
        top_bar_layout.addStretch()

        refresh_btn = QPushButton("Verileri Yenile")
        refresh_btn.setFixedWidth(150)
        refresh_btn.setCursor(Qt.CursorShape.PointingHandCursor)
        refresh_btn.clicked.connect(self.refresh_data)

        top_bar_layout.addWidget(refresh_btn)
        layout.addLayout(top_bar_layout)

        # --- Üst Kartlar ---
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(25)

        # Normal Kart
        self.card_total_stock = self.create_info_card("TOPLAM STOK DEĞERİ", "0.00₺", "#3b82f6")

        # Tıklanabilir Kartlar
        # Index 2: Düşük Stok Uyarıları
        self.card_low_stock = self.create_info_card("KRİTİK STOK", "0", "#ef4444", target_tab_index=2)

        # Index 5: Satışlar
        self.card_monthly_sales = self.create_info_card("BU AY SATIŞ", "0.00₺", "#10b981", target_tab_index=5)

        cards_layout.addWidget(self.card_total_stock)
        cards_layout.addWidget(self.card_low_stock)
        cards_layout.addWidget(self.card_monthly_sales)
        layout.addLayout(cards_layout)

        # --- Grafikler (Üst Satır) ---
        charts_layout = QHBoxLayout()
        charts_layout.setSpacing(25)

        # Sol: Kategori Dağılımı (Bar)
        self.category_chart = SimpleChartWidget("bar", title="Stok Dağılımı (Adet)")
        self.category_chart_frame = self.wrap_chart(self.category_chart)
        charts_layout.addWidget(self.category_chart_frame, 1)

        # Sağ: Kategori Bazlı Değer (Pasta)
        self.value_chart = SimpleChartWidget("pie", title="Kategori Bazlı Stok Değeri")
        self.value_chart_frame = self.wrap_chart(self.value_chart)
        charts_layout.addWidget(self.value_chart_frame, 1)

        layout.addLayout(charts_layout, 2)

        # --- Alt Grafikler (Alt Satır) ---
        bottom_charts_layout = QHBoxLayout()

        # Alt: Satış Trendi (Line)
        self.sales_chart = SimpleChartWidget("line", title="Son 7 Günlük Satış Trendi")
        self.sales_chart_frame = self.wrap_chart(self.sales_chart)
        bottom_charts_layout.addWidget(self.sales_chart_frame)

        layout.addLayout(bottom_charts_layout, 2)

        # --- Sinyal Bağlantıları ---
        self.sales_chart.clicked.connect(lambda: self.open_chart_detail(self.sales_chart))
        self.category_chart.clicked.connect(lambda: self.open_chart_detail(self.category_chart))
        self.value_chart.clicked.connect(lambda: self.open_chart_detail(self.value_chart))

        # --- Alt Bilgi ---

        # Widget'ı ScrollArea'ya ata
        scroll_area.setWidget(content_widget)

        # ScrollArea'yı ana layout'a ekle
        main_layout.addWidget(scroll_area)

    def create_info_card(self, title, value, color_code, target_tab_index=None):
        if target_tab_index is not None:
            frame = ClickableCard()
            frame.clicked.connect(lambda: self.navigation_requested.emit(target_tab_index))
            frame.setProperty("clickable", True)
        else:
            frame = QFrame()
            frame.setProperty("clickable", False)

        frame.setObjectName("metricCard")
        # Direct QSS injection for dynamic border color, cleaner than f-string in main block
        frame.setStyleSheet(f"#metricCard {{ border-left: 4px solid {color_code}; }}")

        l = QVBoxLayout(frame)
        l.setContentsMargins(20, 20, 20, 20)
        title_lbl = QLabel(title)
        title_lbl.setObjectName("metricTitle")
        val_lbl = QLabel(value)
        val_lbl.setObjectName("metricValue")
        l.addWidget(title_lbl)
        l.addWidget(val_lbl)
        frame.val_lbl = val_lbl

        # --- SHADOW EFFECT ---
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(25)
        shadow.setXOffset(0)
        shadow.setYOffset(8)
        shadow.setColor(QColor(0, 0, 0, 60))  # More subtle, professional shadow
        frame.setGraphicsEffect(shadow)

        return frame

    def wrap_chart(self, chart_widget):
        frame = QFrame()
        frame.setObjectName("chartCard")
        l = QVBoxLayout(frame)
        l.setContentsMargins(5, 5, 5, 5)  # Slight padding for the chart content
        l.addWidget(chart_widget)

        # --- SHADOW EFFECT ---
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(25)
        shadow.setXOffset(0)
        shadow.setYOffset(8)
        shadow.setColor(QColor(0, 0, 0, 60))
        frame.setGraphicsEffect(shadow)

        return frame

    def open_chart_detail(self, chart_widget):
        dialog = ChartDetailDialog(chart_widget.chart_type, chart_widget.title, chart_widget.data,
                                   is_dark=chart_widget.is_dark, parent=self)
        dialog.exec()

    def refresh_data(self):
        """Tüm grafik ve kart verilerini veritabanından tazeleyerek yeniden çizer."""
        try:
            # 1. Kart Verileri
            genel = self.veritabani.genel_bakis_getir()
            self.card_total_stock.val_lbl.setText(f"{genel['toplam_deger']:,.2f}₺")
            self.card_low_stock.val_lbl.setText(str(genel['dusuk_stok']))

            # Aylık Satış
            rapor = self.veritabani.kar_zarar_raporu_getir('aylik')
            aylik_ciro = sum([r[6] for r in rapor]) if rapor else 0.0  # 6. index toplam ciro
            self.card_monthly_sales.val_lbl.setText(f"{aylik_ciro:,.2f}₺")

            # 2. Grafik Verileri
            dates = []
            sales = []
            cursor = self.veritabani.baglanti.cursor()
            for i in range(6, -1, -1):
                d = QDate.currentDate().addDays(-i)
                d_str = d.toString("yyyy-MM-dd")
                # O günkü toplam satış cirosu
                cursor.execute("""
                    SELECT SUM(ABS(miktar_degisimi) * satis_fiyati) 
                    FROM stok_hareketleri 
                    WHERE islem_tipi = 'STOK ÇIKIŞI' AND date(tarih) = ?
                """, (d_str,))
                row = cursor.fetchone()
                res = row[0] if row else 0.0
                sales.append(res if res else 0.0)
                dates.append(QLocale(QLocale.Language.Turkish).toString(d, "dd MMM"))

            self.sales_chart.data = dict(zip(dates, sales))
            self.sales_chart.update()

            # Kategori Dağılımı
            cursor.execute(
                "SELECT kategori, SUM(miktar) FROM urunler GROUP BY kategori ORDER BY SUM(miktar) DESC LIMIT 5")
            cats = cursor.fetchall()
            cat_data = {}
            if cats:
                for c in cats:
                    cat_name = c[0] if c[0] else "Diğer"
                    cat_qty = c[1] if c[1] else 0
                    if cat_qty > 0:  # Sadece 0'dan büyükleri ekle
                        cat_data[cat_name] = cat_qty

            self.category_chart.data = cat_data
            self.category_chart.update()

            # Kategori Bazlı Değer (Pie Chart)
            cursor.execute(
                "SELECT kategori, SUM(miktar * fiyat) FROM urunler GROUP BY kategori ORDER BY SUM(miktar * fiyat) DESC LIMIT 5")
            val_cats = cursor.fetchall()
            val_data = {}
            if val_cats:
                for c in val_cats:
                    cat_name = c[0] if c[0] else "Diğer"
                    cat_val = c[1] if c[1] else 0
                    if cat_val > 0:  # Sadece 0'dan büyükleri ekle
                        val_data[cat_name] = cat_val

            self.value_chart.data = val_data
            self.value_chart.update()



        except Exception as e:
            print(f"Dashboard refresh error: {e}")
            # Hata olsa bile uygulamanın çökmemesi için sessizce devam et veya logla
            # Kullanıcıya göstermek istersen:
            # QMessageBox.warning(self, "Dashboard Hatası", str(e))

# =============================================================================
# İNTERAKTİF TUR (ONBOARDING)
# =============================================================================

class TourOverlay(QWidget):
    """
    Ekranı karartıp belirli bir widget'ı vurgulayan (spotlight) ve
    yanında açıklama kutusu gösteren modern interaktif tur katmanı.
    """

    # Animasyon için özel özellik tanımlıyoruz
    def get_spotlight_rect(self):
        return self._current_spotlight_rect

    def set_spotlight_rect(self, rect):
        self._current_spotlight_rect = rect
        self.update()  # Her karede yeniden çiz

    # PyQt Property tanımı (Animation Framework için)
    spotlight_rect = pyqtProperty(QRectF, get_spotlight_rect, set_spotlight_rect)

    def __init__(self, parent_window, steps):
        super().__init__(parent_window)
        self.parent_window = parent_window
        self.steps = steps
        self.current_step_index = 0

        # Animasyon Başlangıç Değeri
        self._current_spotlight_rect = QRectF(self.rect().center().x(), self.rect().center().y(), 0, 0)

        # --- ANİMASYON GRUBU ---
        self.anim_group = QParallelAnimationGroup(self)

        # 1. Spotlight Animasyonu
        self.anim_spotlight = QPropertyAnimation(self, b"spotlight_rect")
        self.anim_spotlight.setDuration(600)
        self.anim_spotlight.setEasingCurve(QEasingCurve.Type.InOutCubic)

        self.anim_group.addAnimation(self.anim_spotlight)

        # Tam ekran kapla
        self.setGeometry(parent_window.rect())
        self.setAttribute(Qt.WidgetAttribute.WA_TransparentForMouseEvents, False)  # Fareyi yakala
        self.setMouseTracking(True)

        # --- BİLGİ KUTUSU (INFO BOX) ---
        self.info_box = QFrame(self)
        self.info_box.setObjectName("tourInfoBox")

        # 2. Kutu Pozisyon Animasyonu
        self.anim_box = QPropertyAnimation(self.info_box, b"pos")
        self.anim_box.setDuration(600)
        self.anim_box.setEasingCurve(QEasingCurve.Type.InOutCubic)

        self.anim_group.addAnimation(self.anim_box)

        # Gölge Efekti
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(25)
        shadow.setXOffset(0)
        shadow.setYOffset(8)
        shadow.setColor(QColor(0, 0, 0, 80))
        self.info_box.setGraphicsEffect(shadow)

        # Stil Tanımları (QSS)
        self.info_box.setStyleSheet("""
            QFrame#tourInfoBox { 
                background-color: #ffffff; 
                border-radius: 12px; 
                border: 1px solid #e2e8f0;
            }
            QLabel#tourTitle { 
                color: #0f172a; 
                font-weight: 800; 
                font-size: 18px; 
                margin-bottom: 5px;
            }
            QLabel#tourDesc { 
                color: #475569; 
                font-size: 14px; 
                line-height: 1.4;
            }
            QLabel#tourCounter {
                color: #94a3b8;
                font-size: 11px;
                font-weight: 600;
            }
            QPushButton { 
                border-radius: 6px; 
                font-weight: 600;
                font-size: 13px;
                padding: 8px 16px;
            }
            QPushButton#nextBtn {
                background-color: #3b82f6; 
                color: white; 
                border: none;
            }
            QPushButton#nextBtn:hover { background-color: #2563eb; }
            QPushButton#nextBtn:pressed { background-color: #1d4ed8; }

            QPushButton#prevBtn {
                background-color: transparent; 
                color: #64748b; 
                border: 1px solid transparent;
            }
            QPushButton#prevBtn:hover { color: #334155; background-color: #f1f5f9; }

            QPushButton#closeBtn {
                background-color: transparent;
                color: #cbd5e1;
                border: none;
                font-size: 16px;
                font-weight: bold;
            }
            QPushButton#closeBtn:hover { color: #ef4444; }
        """)

        # Layout
        self.box_layout = QVBoxLayout(self.info_box)
        self.box_layout.setContentsMargins(24, 20, 24, 24)
        self.box_layout.setSpacing(12)

        # Üst Kısım
        top_layout = QHBoxLayout()
        self.lbl_counter = QLabel("ADIM 1 / 5")
        self.lbl_counter.setObjectName("tourCounter")

        self.btn_close_top = QPushButton("×")
        self.btn_close_top.setObjectName("closeBtn")
        self.btn_close_top.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_close_top.setFixedSize(30, 30)
        self.btn_close_top.clicked.connect(self.close_tour)

        top_layout.addWidget(self.lbl_counter)
        top_layout.addStretch()
        top_layout.addWidget(self.btn_close_top)

        # İçerik
        self.lbl_title = QLabel()
        self.lbl_title.setObjectName("tourTitle")
        self.lbl_title.setWordWrap(True)

        self.lbl_desc = QLabel()
        self.lbl_desc.setObjectName("tourDesc")
        self.lbl_desc.setWordWrap(True)
        self.lbl_desc.setMinimumHeight(60)

        # Alt Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)

        self.btn_prev = QPushButton("Geri")
        self.btn_prev.setObjectName("prevBtn")
        self.btn_prev.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_prev.clicked.connect(self.prev_step)

        self.btn_next = QPushButton("İlerle")
        self.btn_next.setObjectName("nextBtn")
        self.btn_next.setCursor(Qt.CursorShape.PointingHandCursor)
        self.btn_next.clicked.connect(self.next_step)

        btn_layout.addWidget(self.btn_prev)
        btn_layout.addStretch()
        btn_layout.addWidget(self.btn_next)

        # Hepsini Birleştir
        self.box_layout.addLayout(top_layout)
        self.box_layout.addWidget(self.lbl_title)
        self.box_layout.addWidget(self.lbl_desc)
        self.box_layout.addStretch()
        self.box_layout.addLayout(btn_layout)

        self.info_box.setFixedWidth(360)
        # Yüksekliği içeriğe göre ayarlasın ama min biraz olsun

        # İlk adımı yükle (Biraz gecikmeli ki layout otursun)
        self.center_info_box()  # Başlangıçta ortada olsun
        from PyQt6.QtCore import QTimer
        QTimer.singleShot(100, self.show_step)
        self.show()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        # 1. Tam ekran yarı saydam arka plan
        painter.setBrush(QColor(15, 23, 42, 180))
        painter.setPen(Qt.PenStyle.NoPen)

        full_region = self.rect()

        # --- SPOTLIGHT EFEKTİ (Hole) ---
        focus_rect = self._current_spotlight_rect

        # Eğer rect çok küçükse
        if focus_rect.width() < 5 or focus_rect.height() < 5:
            painter.drawRect(full_region)
            return

        path = QPainterPath()
        path.addRect(QRectF(full_region))
        path.addRoundedRect(focus_rect, 10, 10)  # Köşeler yuvarlatılmış
        path.setFillRule(Qt.FillRule.OddEvenFill)
        painter.drawPath(path)

        # --- VURGU ÇERÇEVESİ ---
        painter.setBrush(Qt.BrushStyle.NoBrush)

        # Glow efekti için kalın, yarı saydam kalem
        glow_pen = QPen(QColor(59, 130, 246, 120), 6)  # Mavi glow
        painter.setPen(glow_pen)
        painter.drawRoundedRect(focus_rect, 10, 10)

        # Keskin iç çizgi
        sharp_pen = QPen(QColor(59, 130, 246), 2)
        painter.setPen(sharp_pen)
        painter.drawRoundedRect(focus_rect, 10, 10)

    def calculate_box_position(self, target_rect):
        """Kutuyu akıllıca konumlandır"""
        box_w = self.info_box.width()
        box_h = self.info_box.height()
        margin = 20

        # Ekran boyutları
        sw = self.width()
        sh = self.height()

        # Varsayılan: Hedefin SAĞINDA
        x = target_rect.right() + margin
        y = target_rect.top()

        # Sağa sığmıyorsa SOLA al
        if x + box_w > sw - margin:
            x = target_rect.left() - box_w - margin

        # Sola da sığmıyorsa (mobil/dar ekran) veya çok dar alan varsa AŞAĞI al
        if x < margin or (target_rect.width() > sw * 0.6):
            x = max(margin, target_rect.left())  # Hizala
            y = target_rect.bottom() + margin

            # Aşağı sığmıyorsa YUKARI al (veya ortala)
            if y + box_h > sh - margin:
                y = target_rect.top() - box_h - margin

        # Dikey taşma kontrolü (Ekran dışına çıkmasın)
        if y + box_h > sh - margin:
            y = sh - box_h - margin
        if y < margin:
            y = margin

        # Yatay taşma kontrolü
        if x + box_w > sw - margin:
            x = sw - box_w - margin
        if x < margin:
            x = margin

        return QPoint(int(x), int(y))

    def center_info_box(self):
        # Bu da artık animasyonla olabilir ama şimdilik direct set
        target_pos = QPoint(
            (self.width() - self.info_box.width()) // 2,
            (self.height() - self.info_box.height()) // 2
        )
        self.info_box.move(target_pos)

    def show_step(self):
        step = self.steps[self.current_step_index]
        self.lbl_title.setText(step['title'])
        self.lbl_desc.setText(step['desc'])

        # Sayaç Güncelle
        total = len(self.steps)
        self.lbl_counter.setText(f"ADIM {self.current_step_index + 1} / {total}")

        self.info_box.adjustSize()

        # Sayfa Değişimi Kontrolü
        page_changed = False
        if 'page_index' in step:
            current_row = self.parent_window.nav_list.currentRow()
            if current_row != step['page_index']:
                self.parent_window.nav_list.setCurrentRow(step['page_index'])
                page_changed = True

        # HEDEF HESAPLAMA VE ANİMASYON BAŞLATMA
        # Layout'un oturması için eğer sayfa değiştiyse biraz beklememiz gerekebilir.
        # Bu yüzden asıl hesaplamayı bir sonraki event loop'a veya QTimer'a bırakıyoruz.

        from PyQt6.QtCore import QTimer
        # Eğer sayfa değiştiyse 150ms bekle, değişmediyse hemen (10ms)
        delay = 150 if page_changed else 10
        QTimer.singleShot(delay, self._animate_to_target)

        # Buton Durumları
        self.btn_prev.setEnabled(self.current_step_index > 0)
        self.btn_prev.setVisible(self.current_step_index > 0)

        if self.current_step_index == len(self.steps) - 1:
            self.btn_next.setText("Turu Tamamla")
            self.btn_next.setStyleSheet("""
                QPushButton#nextBtn {
                    background-color: #10b981; color: white; border: none;
                }
                QPushButton#nextBtn:hover { background-color: #059669; }
            """)
        else:
            self.btn_next.setText("Sıradaki")
            self.btn_next.setStyleSheet("""
                QPushButton#nextBtn {
                   background-color: #3b82f6; color: white; border: none;
                }
                QPushButton#nextBtn:hover { background-color: #2563eb; }
            """)

    def _animate_to_target(self):
        """Asıl hedefi bulup animasyon grubunu başlatır"""
        step = self.steps[self.current_step_index]
        target_widget = step.get('widget')

        # Başlangıç değerleri (Mevcut durum)
        start_rect = self._current_spotlight_rect
        start_pos = self.info_box.pos()

        # Bitiş değerlerini hesapla
        end_rect = QRectF(self.rect().center().x(), self.rect().center().y(), 0, 0)  # Fallback: Center (hidden)
        end_pos = QPoint(
            (self.width() - self.info_box.width()) // 2,
            (self.height() - self.info_box.height()) // 2
        )

        if target_widget and target_widget.isVisible():
            try:
                # Koordinat hesapla
                global_pos = target_widget.mapToGlobal(QPoint(0, 0))
                local_pos = self.mapFromGlobal(global_pos)

                # Hedef rect
                w = target_widget.width()
                h = target_widget.height()
                padding = 6

                # Eğer widget çok küçükse veya hatalıysa kontrol et
                if w > 0 and h > 0:
                    end_rect = QRectF(local_pos.x(), local_pos.y(), w, h).adjusted(-padding, -padding, padding, padding)
                    end_pos = self.calculate_box_position(end_rect)
            except RuntimeError:
                # Widget silinmiş olabilir
                pass

        # Animasyonları ayarla
        self.anim_group.stop()

        self.anim_spotlight.setStartValue(start_rect)
        self.anim_spotlight.setEndValue(end_rect)

        self.anim_box.setStartValue(start_pos)
        self.anim_box.setEndValue(end_pos)

        self.anim_group.start()

    def next_step(self):
        if self.current_step_index < len(self.steps) - 1:
            self.current_step_index += 1
            self.show_step()
        else:
            self.close_tour()

    def prev_step(self):
        if self.current_step_index > 0:
            self.current_step_index -= 1
            self.show_step()

    def close_tour(self):
        self.close()
        self.deleteLater()

    def resizeEvent(self, event):
        # Pencere boyutu değişirse overlay'i de güncelle
        self.setGeometry(self.parent_window.rect())
        super().resizeEvent(event)


# =============================================================================
# 6. MAIN APPLICATION WINDOW
# =============================================================================

class AnaPencere(QMainWindow):
    """
    Uygulamanın ana çerçevesi (Main Window).
    Sol taraftaki sidebar ve sağ taraftaki içerik alanını (Stacked Widget) yönetir.
    Kullanıcı etkileşimlerini (Menü, Sayfa Geçişleri, Dialoglar) koordine eder.
    """
    cikis_istendi = pyqtSignal()

    def __init__(self, kullanici_adi, veritabani_yoneticisi):
        super().__init__()
        self.kullanici_adi = kullanici_adi
        self.veritabani = veritabani_yoneticisi
        self.setWindowTitle(f"StockFlow - {kullanici_adi}")
        self.setGeometry(100, 100, 1400, 900)
        self.setMinimumSize(1200, 700)
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.main_layout = QHBoxLayout(self.central_widget)
        self.main_layout.setContentsMargins(0, 0, 0, 0)
        self.main_layout.setSpacing(0)
        self.sayfa_gecmisi = []  # Ziyaret edilen sayfaların listesi
        self.aktif_sayfa_index = 0  # Şu anki sayfa
        self.geri_basildi_flag = False
        self.koyu_tema_aktif = False
        self.init_ui()
        self.init_menu_actions()

    def init_ui(self):
        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebarFrame")
        self.sidebar.setFixedWidth(256)
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(12, 12, 12, 12)

        # --- Logo Kısmı ---
        logo_icon = QLabel()
        logo_icon.setObjectName("logoIcon")
        try:
            logo_path = dosya_yolunu_bul("StockFlow_Logo.png")
            logo_pixmap = QPixmap(logo_path)
            if not logo_pixmap.isNull():
                logo_icon.setPixmap(logo_pixmap.scaled(200, 300, Qt.AspectRatioMode.KeepAspectRatio,
                                                       Qt.TransformationMode.SmoothTransformation))
                logo_icon.setMinimumHeight(50)
            else:
                raise FileNotFoundError
        except FileNotFoundError:
            # Emoji kaldırıldı, yerine metin
            logo_icon.setText("StockFlow")
            logo_icon.setStyleSheet("font-size: 24px; font-weight: bold; color: #94a3b8; padding: 10px; margin: 0;")

        logo_layout = QHBoxLayout()
        logo_layout.addWidget(logo_icon)
        logo_layout.addStretch()

        # --- Navigasyon Listesi ---
        self.nav_list = QListWidget()
        self.nav_list.setObjectName("sidebarNav")

        # Öğeleri ekle
        self.nav_list.addItem(QListWidgetItem("Genel Bakış (Dashboard)"))
        self.nav_list.addItem(QListWidgetItem("Ana Stok Listesi"))
        self.nav_list.addItem(QListWidgetItem("Düşük Stok Uyarıları"))
        self.nav_list.addItem(QListWidgetItem("Stok Hareketleri"))
        self.nav_list.addItem(QListWidgetItem("En Çok Satanlar"))
        self.nav_list.addItem(QListWidgetItem("Satışlar"))

        # --- YENİ EKLENEN AYARLAR ---
        # 1. Dikey kaydırma çubuğunu otomatik yapıyoruz
        self.nav_list.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)
        # 2. Yatay kaydırma çubuğunu kapatıyoruz
        self.nav_list.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        # 3. Listenin tüm öğeleri alacak kadar yüksek olmasını sağlıyoruz (Piksell cinsinden)
        self.nav_list.setMinimumHeight(350)
        # 4. Seçim yapıldığında mavi çerçevenin çıkmasını engellemek için (görsel iyileştirme)
        self.nav_list.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        # ----------------------------

        source_button = QPushButton(f"Kullanıcı: {self.kullanici_adi}")
        source_button.setObjectName("sourceButton")
        source_button.setEnabled(False)

        sidebar_layout.addLayout(logo_layout)
        sidebar_layout.addWidget(QLabel("NAVİGASYON"), 0, Qt.AlignmentFlag.AlignLeft)

        # Listeyi ekle
        sidebar_layout.addWidget(self.nav_list)

        # Listenin altına boşluk ekleyerek butonu en alta itiyoruz
        sidebar_layout.addStretch()

        # --- TEMA DEĞİŞTİRME BUTONU ---
        container_toggle = QWidget()
        layout_toggle = QHBoxLayout(container_toggle)
        layout_toggle.setContentsMargins(0, 0, 0, 10)  # Kenar boşluklarını sıfırladık
        layout_toggle.setAlignment(Qt.AlignmentFlag.AlignCenter)  # Ortala

        self.theme_toggle = ThemeToggleButton(self, self.koyu_tema_aktif)
        self.theme_toggle.toggled.connect(self.tema_degistir_toggle)

        layout_toggle.addWidget(self.theme_toggle)
        # -------------------------------------

        # Tema butonunu ekle
        sidebar_layout.addWidget(container_toggle)

        sidebar_layout.addWidget(source_button)

        self.main_layout.addWidget(self.sidebar)

        # --- Buradan sonrası aynı (Header ve Content) ---
        content_container = QFrame()
        content_container.setObjectName("contentContainer")
        content_layout = QVBoxLayout(content_container)
        content_layout.setContentsMargins(0, 0, 0, 0)
        content_layout.setSpacing(0)
        header_bar = QFrame()
        header_bar.setObjectName("headerBar")
        header_bar_layout = QHBoxLayout(header_bar)
        header_bar_layout.setContentsMargins(20, 10, 20, 10)
        self.geri_btn = QPushButton("Geri")
        self.geri_btn.setObjectName("headerButton")
        self.geri_btn.setFixedWidth(100)
        self.geri_btn.clicked.connect(self.geri_git)
        self.sayfa_basligi = QLabel("Genel Bakış")
        self.sayfa_basligi.setObjectName("pageTitle")
        self.ayarlar_btn = QPushButton("Ayarlar")
        self.ayarlar_btn.setObjectName("headerButton")
        self.ayarlar_btn.setFixedWidth(100)
        self.ayarlar_btn.clicked.connect(self.ayarlar_menu_goster)
        header_bar_layout.addWidget(self.geri_btn)
        header_bar_layout.addStretch()
        header_bar_layout.addWidget(self.sayfa_basligi)
        header_bar_layout.addStretch()
        header_bar_layout.addWidget(self.ayarlar_btn)
        content_layout.addWidget(header_bar)
        self.stacked_widget = QStackedWidget()
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)

        # Sayfalar
        self.dashboard_sayfasi = DashboardPage(self.veritabani)
        # --- Dashboard'dan gelen navigasyon isteğini karşıla ---
        self.dashboard_sayfasi.navigation_requested.connect(self.nav_list.setCurrentRow)

        self.ana_stok_sayfasi = AnaStokSayfasi(self.veritabani, self.status_bar, self.kullanici_adi)
        self.dusuk_stok_sayfasi = DusukStokSayfasi(self.veritabani)
        self.gecmis_sayfasi = StokHareketSayfasi(self.veritabani, self.kullanici_adi)
        self.satis_raporu_sayfasi = SatisRaporuSayfasi(self.veritabani)
        self.kar_zarar_sayfasi = KarZararSayfasi(self.veritabani)

        self.stacked_widget.addWidget(self.dashboard_sayfasi)  # Index 0
        self.stacked_widget.addWidget(self.ana_stok_sayfasi)  # Index 1
        self.stacked_widget.addWidget(self.dusuk_stok_sayfasi)  # Index 2
        self.stacked_widget.addWidget(self.gecmis_sayfasi)  # Index 3
        self.stacked_widget.addWidget(self.satis_raporu_sayfasi)  # Index 4
        self.stacked_widget.addWidget(self.kar_zarar_sayfasi)  # Index 5

        content_layout.addWidget(self.stacked_widget, 1)
        self.main_layout.addWidget(content_container, 1)
        self.nav_list.currentRowChanged.connect(self.sayfa_degisti)
        self.nav_list.setCurrentRow(0)
        self.status_bar.showMessage(f"Hoş geldiniz, {self.kullanici_adi}!", 5000)

    def init_menu_actions(self):
        # Menü nesnesini oluştur
        self.ayarlar_menu = QMenu(self)

        # 0. Yardım ve Hakkında
        self.kilavuz_action = QAction("Kullanım Kılavuzu", self)
        self.kilavuz_action.triggered.connect(self.kullanim_kilavuzu_ac)

        self.hakkinda_action = QAction("Uygulama Hakkında", self)
        self.hakkinda_action.triggered.connect(self.hakkinda_dialogu_ac)

        # 1. Kullanıcı İşlemleri
        self.kullanici_degistir_action = QAction("Mevcut Kullanıcı Bilgilerini Değiştir", self)
        self.kullanici_degistir_action.triggered.connect(self.kullanici_degistir_dialogu_ac)

        self.yeni_kullanici_action = QAction("Yeni Kullanıcı Ekle", self)
        self.yeni_kullanici_action.triggered.connect(self.yeni_kullanici_dialogu_ac)

        # 2. Firebase ve Dışa Aktarma
        self.firebase_action = QAction("Firebase Senkronizasyon", self)
        self.firebase_action.triggered.connect(self.firebase_penceresi_ac)

        self.disa_aktar_action = QAction("Verileri Dışa Aktar (Excel)", self)
        self.disa_aktar_action.triggered.connect(self.detayli_excel_raporu_olustur)

        # 3. VERİ SIFIRLAMA
        self.sifirlama_action = QAction("Veri Tabanını Sıfırla / Temizle", self)
        self.sifirlama_action.triggered.connect(self.veri_sifirlama_islemi)

        # 4. Çıkış
        self.cikis_action = QAction("Çıkış Yap", self)
        self.cikis_action.triggered.connect(self.oturumu_kapat)

        # --- AKSİYONLARI MENÜYE EKLEME ---
        self.ayarlar_menu.addAction(self.kilavuz_action)  # En üste
        self.ayarlar_menu.addAction(self.hakkinda_action)
        self.ayarlar_menu.addSeparator()

        self.ayarlar_menu.addAction(self.kullanici_degistir_action)
        self.ayarlar_menu.addAction(self.yeni_kullanici_action)
        self.ayarlar_menu.addSeparator()

        self.ayarlar_menu.addAction(self.firebase_action)
        self.ayarlar_menu.addAction(self.disa_aktar_action)
        self.ayarlar_menu.addSeparator()

        self.ayarlar_menu.addAction(self.sifirlama_action)

        self.ayarlar_menu.addSeparator()
        self.ayarlar_menu.addAction(self.cikis_action)

    def tema_degistir(self):
        """Menüden veya başka yerden tetiklenen tema değişimi"""
        self.koyu_tema_aktif = not self.koyu_tema_aktif

        # Toggle butonu varsa onu da güncelle
        if hasattr(self, 'theme_toggle'):
            self.theme_toggle._is_dark = self.koyu_tema_aktif
            self.theme_toggle.position = 1.0 if self.koyu_tema_aktif else 0.0

        self.uygula_tema_degisikligi()

    def tema_degistir_toggle(self, is_dark):
        """Toggle butonundan gelen sinyalle temayı değiştirir."""
        self.koyu_tema_aktif = is_dark
        self.uygula_tema_degisikligi()

    def uygula_tema_degisikligi(self):
        """
        Seçilen temayı (Açık/Koyu) tüm uygulamaya uygular.
        Geçiş sırasında ekran görüntüsü alarak 'fade' animasyonu yapar.
        """
        # --- ANIMASYON BAŞLANGICI ---
        # 1. Mevcut ekranın görüntüsünü al
        try:
            # Pencerenin o anki halini pixmap olarak yakala
            screenshot = self.grab()

            # 2. Üzerine geçici bir etiket (overlay) oluştur
            self.fade_overlay = QLabel(self)
            self.fade_overlay.setPixmap(screenshot)
            self.fade_overlay.setGeometry(self.rect())
            self.fade_overlay.show()

            # 3. Saydamlık Efekti
            self.fade_effect = QGraphicsOpacityEffect(self.fade_overlay)
            self.fade_overlay.setGraphicsEffect(self.fade_effect)

            # 4. Animasyon Tanımı (1.0 -> 0.0 Opacity)
            self.fade_anim = QPropertyAnimation(self.fade_effect, b"opacity")
            self.fade_anim.setDuration(400)  # 400ms sürsün
            self.fade_anim.setStartValue(1.0)
            self.fade_anim.setEndValue(0.0)
            self.fade_anim.setEasingCurve(QEasingCurve.Type.OutQuad)

            # 5. Animasyon bitince temizle
            self.fade_anim.finished.connect(self.fade_overlay.deleteLater)
            self.fade_anim.start()

        except Exception as e:
            print(f"Animasyon hatası: {e}")
        # ----------------------------

        # 1. Yeni Stili Uygula (Arka planda stil değişir, overlay yavaşça silinirken yeni stil görünür)
        yeni_stil = stil_olustur(self.koyu_tema_aktif)
        QApplication.instance().setStyleSheet(yeni_stil)

        # 2. Menü Yazısını Güncelle (Eğer menüde varsa)
        if hasattr(self, 'tema_action'):
            if self.koyu_tema_aktif:
                self.tema_action.setText("Açık Temaya Geç")
            else:
                self.tema_action.setText("Koyu Temaya Geç")

        # 3. Grafiklerin Renklerini Güncelle
        # Dashboard'daki grafiklere ulaşalım
        charts = [
            self.dashboard_sayfasi.sales_chart,
            self.dashboard_sayfasi.category_chart,
            self.dashboard_sayfasi.value_chart
        ]

        for chart in charts:
            chart.set_theme(self.koyu_tema_aktif)

    def kullanim_kilavuzu_ac(self):
        # Tur Adımları Tanımlama
        steps = [
            {
                'title': 'StockFlow\'a Hoş Geldiniz!',
                'desc': 'Uygulamamızın temel özelliklerini hızlıca keşfetmeye hazır mısınız? Bu kısa tur size rehberlik edecek.',
                'widget': None,  # Ortada göster
                'page_index': 0  # Dashboard'a git
            },
            {
                'title': 'Navigasyon Menüsü',
                'desc': 'Sol taraftaki bu menüyü kullanarak stok listesi, raporlar ve satış ekranları arasında geçiş yapabilirsiniz.',
                'widget': self.sidebar,
                'page_index': 0
            },
            {
                'title': 'Anlık Durum Kartları',
                'desc': 'İşletmenizin o anki stok değeri ve kritik uyarıları burada özetlenir.\n\nİpucu: Kırmızı ve yeşil kartlara tıklayarak detaylara gidebilirsiniz.',
                'widget': self.dashboard_sayfasi.findChild(QWidget, "metricCard"),
                # İlk kartı bulmaya çalışır (Toplam Stok)
                'page_index': 0
            },
            {
                'title': 'Grafiksel Analiz',
                'desc': 'Satış trendlerini ve stok dağılımını bu grafiklerden takip edebilirsiniz. Detaylı analiz için grafiklerin üzerine tıklayın.',
                'widget': self.dashboard_sayfasi.sales_chart_frame,
                'page_index': 0
            },
            {
                'title': 'Stok Yönetimi',
                'desc': 'Tüm ürünlerinizi buradan yönetirsiniz. Yeni ürün eklemek için (+), düzenlemek veya satış yapmak için (•••) butonunu kullanın.',
                'widget': self.ana_stok_sayfasi.stok_tablosu,  # Tabloyu vurgula
                'page_index': 1  # Stok Sayfasına Git
            },
            {
                'title': 'Düşük Stok Uyarıları',
                'desc': 'Stok miktarı, belirlediğiniz kritik seviyenin altına düşen ürünler burada listelenir. Eksikleri buradan takip edip sipariş verebilirsiniz.',
                'widget': self.dusuk_stok_sayfasi.stok_tablosu,
                'page_index': 2  # Düşük Stok Sayfasına Git
            },
            {
                'title': 'Stok Geçmişi & Loglar',
                'desc': 'Hangi ürün ne zaman eklendi, ne zaman satıldı? Tüm giriş-çıkış hareketleri ve kullanıcı işlemleri burada kayıtlıdır.',
                'widget': self.gecmis_sayfasi.rapor_tablosu,
                'page_index': 3  # Geçmiş Sayfasına Git
            },
            {
                'title': 'Satış Raporları',
                'desc': 'Belirli tarih aralıklarındaki satış performansınızı ve ciro detaylarınızı bu ekrandan analiz edebilirsiniz.',
                'widget': self.satis_raporu_sayfasi.tablo,
                'page_index': 4  # Satış Raporu Sayfasına Git
            },
            {
                'title': 'Ayarlar Menüsü',
                'desc': 'Kullanıcı işlemleri, yedekleme ve veri sıfırlama gibi yönetimsel araçlara buradan ulaşabilirsiniz.',
                'widget': self.ayarlar_btn,
                'page_index': 0  # Dashboard'a geri dön
            }
        ]

        self.tour = TourOverlay(self, steps)

    def hakkinda_dialogu_ac(self):
        QMessageBox.information(self, "StockFlow Hakkında",
                                "<b>StockFlow Stok Takip Sistemi</b><br>"
                                "Sürüm: 2.1.0<br>"
                                "Geliştirici: StockFlow Team<br><br>"
                                "Modern, hızlı ve güvenilir stok yönetimi.<br>"
                                "Tüm hakları saklıdır © 2025")

    def oturumu_kapat(self):
        """Mevcut oturumu kapatır ve kontrolcüye sinyal göndererek giriş ekranına dönülmesini sağlar."""
        self.cikis_istendi.emit()  # Kontrolcüye haber ver (Giriş ekranını açar)
        self.close()

    def ayarlar_menu_goster(self):
        self.ayarlar_menu.exec(self.ayarlar_btn.mapToGlobal(QPoint(0, self.ayarlar_btn.height())))

    def kullanici_degistir_dialogu_ac(self):
        dialog = KullaniciDegistirPenceresi(self.veritabani, self.kullanici_adi)
        dialog.exec()

    def yeni_kullanici_dialogu_ac(self):
        dialog = YeniKullaniciDialog(self.veritabani, None)
        dialog.exec()

    def firebase_penceresi_ac(self):
        dialog = FirebaseYedekleyici(self.veritabani, self)
        dialog.exec()
        # Veri çekildiyse arayüzü yenile
        self.ana_stok_sayfasi.stogu_guncelle_arayuz()

    def geri_git(self):
        """Geçmiş listesindeki son sayfaya döner."""
        if self.sayfa_gecmisi:
            # Listeden son gezilen sayfayı alıyoruz (pop işlemi silerek alır)
            onceki_index = self.sayfa_gecmisi.pop()

            # Geri tuşuna basıldığını işaretliyoruz ki bu geçişi tekrar geçmişe eklemesin
            self.geri_basildi_flag = True

            # Sayfayı değiştir
            self.nav_list.setCurrentRow(onceki_index)

        # Eğer geçmiş bittiyse butonu pasif yapabiliriz (Opsiyonel)
        if not self.sayfa_gecmisi:
            self.geri_btn.setEnabled(False)

    def sayfa_degisti(self, index):
        if not self.geri_basildi_flag:
            # Eğer geri butonuna basılmadıysa (normal tıklama ise)
            # Şu anki sayfayı (değişmeden önceki halini) geçmişe ekle
            if self.aktif_sayfa_index != index:  # Aynı sayfaya tıklanmadıysa
                self.sayfa_gecmisi.append(self.aktif_sayfa_index)
                self.geri_btn.setEnabled(True)  # Geçmiş oluştuğu için butonu aktif et
        else:
            # Geri butonuna basıldıysa flag'i sıfırla, geçmişe ekleme yapma
            self.geri_basildi_flag = False

            # Aktif sayfayı güncelle
        self.aktif_sayfa_index = index
        self.stacked_widget.setCurrentIndex(index)
        basliklar = ["Genel Bakış", "Ana Stok Listesi", "Düşük Stok Uyarıları", "Stok Hareketleri", "En Çok Satanlar",
                     "Satış Raporu"]
        if 0 <= index < len(basliklar):
            self.sayfa_basligi.setText(basliklar[index])

        # Sayfa değiştiğinde verileri yenile
        if index == 0:
            self.dashboard_sayfasi.refresh_data()
        elif index == 1:
            self.ana_stok_sayfasi.stogu_guncelle_arayuz()
        elif index == 2:
            self.dusuk_stok_sayfasi.stogu_guncelle()
        elif index == 3:
            self.gecmis_sayfasi.raporu_guncelle()
        elif index == 4:
            self.satis_raporu_sayfasi.raporu_guncelle()
        elif index == 5:
            self.kar_zarar_sayfasi.raporu_guncelle()

    def hizli_urun_ekle(self):
        self.nav_list.setCurrentRow(1)  # Stok Sayfası
        self.ana_stok_sayfasi.yeni_urun_ekle()

    def veri_sifirlama_islemi(self):

        # 1. KAPSAM SEÇİMİ PENCERESİ
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("Veri Sıfırlama Seçenekleri")
        msg_box.setText("Hangi verileri silmek istiyorsunuz?")
        msg_box.setInformativeText("DİKKAT: Bu işlem geri alınamaz!")
        msg_box.setIcon(QMessageBox.Icon.Warning)

        # Butonları Türkçe tanımlıyoruz
        btn_gecmis = msg_box.addButton("Sadece Geçmişi Sil", QMessageBox.ButtonRole.ActionRole)
        btn_hepsi = msg_box.addButton("Her Şeyi Sil (Fabrika Ayarları)", QMessageBox.ButtonRole.DestructiveRole)
        btn_iptal = msg_box.addButton("İptal", QMessageBox.ButtonRole.RejectRole)

        msg_box.exec()

        secilen_buton = msg_box.clickedButton()
        if secilen_buton == btn_iptal:
            return

        kapsam = 'tumu' if secilen_buton == btn_hepsi else 'gecmis'
        onay_kelimesi = "SIFIRLA"

        # 2. GÜVENLİK ONAYI PENCERESİ (Input Dialog)
        dialog = QInputDialog(self)
        dialog.setWindowTitle("Güvenlik Onayı")
        dialog.setLabelText(f"Bu işlem kalıcıdır.\nOnaylamak için lütfen '{onay_kelimesi}' yazın:")
        dialog.setTextValue("")

        # OK/Cancel yerine Tamam/İptal
        dialog.setOkButtonText("Tamam")
        dialog.setCancelButtonText("İptal")


        ok = dialog.exec()
        text = dialog.textValue()

        if ok and text == onay_kelimesi:
            basari, mesaj = self.veritabani.veritabanini_sifirla(kapsam, self.kullanici_adi)

            # Sonuç mesaj kutusu
            sonuc_box = QMessageBox(self)
            if basari:
                sonuc_box.setWindowTitle("Başarılı")
                sonuc_box.setIcon(QMessageBox.Icon.Information)
                sonuc_box.setText(mesaj)
                sonuc_box.addButton("Tamam", QMessageBox.ButtonRole.AcceptRole)  # Türkçe buton
                sonuc_box.exec()

                # Arayüzü yenile
                self.dashboard_sayfasi.refresh_data()
                self.ana_stok_sayfasi.stogu_guncelle_arayuz()
                self.dusuk_stok_sayfasi.stogu_guncelle()
                self.gecmis_sayfasi.raporu_guncelle()
                self.satis_raporu_sayfasi.raporu_guncelle()
                self.kar_zarar_sayfasi.raporu_guncelle()
                self.nav_list.setCurrentRow(0)
            else:
                sonuc_box.setWindowTitle("Hata")
                sonuc_box.setIcon(QMessageBox.Icon.Critical)
                sonuc_box.setText(mesaj)
                sonuc_box.addButton("Tamam", QMessageBox.ButtonRole.AcceptRole)  # Türkçe buton
                sonuc_box.exec()

        elif ok:
            # Yanlış kelime girilirse
            uyari_box = QMessageBox(self)
            uyari_box.setWindowTitle("İptal Edildi")
            uyari_box.setIcon(QMessageBox.Icon.Warning)
            uyari_box.setText("Doğrulama kelimesi yanlış girildiği için işlem iptal edildi.")
            uyari_box.addButton("Tamam", QMessageBox.ButtonRole.AcceptRole)
            uyari_box.exec()

    def detayli_excel_raporu_olustur(self):
        if not EXCEL_MUMKUN:
            QMessageBox.warning(self, "Eksik Kütüphane",
                                "Excel raporu için 'openpyxl' gereklidir.\nTerminale 'pip install openpyxl' yazarak kurabilirsiniz.")
            return

        # Dosya kaydetme diyaloğu
        dosya_yolu, _ = QFileDialog.getSaveFileName(self, "Kapsamlı Raporu Kaydet",
                                                    f"StockFlow_Full_Rapor_{date.today()}.xlsx",
                                                    "Excel Dosyası (*.xlsx)")

        if not dosya_yolu:
            return

        try:
            wb = openpyxl.Workbook()


            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # --- YARDIMCI FONKSİYON: Sayfa Oluşturucu ---
            def sayfa_ekle(baslik, veriler, kolon_isimleri):
                ws = wb.create_sheet(title=baslik)
                ws.append(kolon_isimleri)

                # Başlık Stili (Koyu Lacivert Arka Plan, Beyaz Yazı)
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")

                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")

                # Verileri Ekle
                for satir in veriler:
                    ws.append(satir)

                # Sütun Genişliklerini Ayarla
                for column_cells in ws.columns:
                    length = max(len(str(cell.value) if cell.value else "") for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2,
                                                                                    50)  # Max 50 karakter genişlik

            # ---------------------------------------------------------
            # 1. SEKME: ANA STOK LİSTESİ
            # ---------------------------------------------------------
            urunler = self.veritabani.urunleri_getir()
            basliklar_stok = ["ID", "Kod", "Ürün Adı", "Kategori", "Alış Fiyatı", "Miktar", "Birim", "Min. Stok",
                              "Başlangıç", "SKT"]
            sayfa_ekle("Stok Listesi", urunler, basliklar_stok)

            # ---------------------------------------------------------
            # 2. SEKME: KRİTİK (DÜŞÜK) STOK
            # ---------------------------------------------------------
            dusuk_stok = self.veritabani.dusuk_stok_urunleri_getir()
            basliklar_dusuk = ["ID", "Kod", "Ürün Adı", "Kategori", "Fiyat", "Mevcut Miktar", "Birim", "Min. Stok"]
            sayfa_ekle("Kritik Stoklar", dusuk_stok, basliklar_dusuk)

            # ---------------------------------------------------------
            # 3. SEKME: STOK HAREKETLERİ (GEÇMİŞ)
            # ---------------------------------------------------------
            # 'tumu' diyerek tüm geçmişi çekiyoruz
            hareketler = self.veritabani.stok_hareketlerini_getir('tumu')
            basliklar_hareket = ["ID", "Tarih", "Kullanıcı", "Kod", "Ürün Adı", "İşlem", "Değişim", "Son Miktar",
                                 "Birim", "Notlar", "Satış Fiyatı"]
            sayfa_ekle("İşlem Geçmişi", hareketler, basliklar_hareket)

            # ---------------------------------------------------------
            # 4. SEKME: SATIŞ RAPORU (FİNANSAL)
            # ---------------------------------------------------------
            # 'tumu' diyerek tüm zamanların finansal raporunu çekiyoruz
            finansal = self.veritabani.kar_zarar_raporu_getir('tumu')
            basliklar_finans = ["Kod", "Ürün Adı", "Kategori", "Satılan Adet", "Alış Fiyatı", "Ort. Satış Fiyatı",
                                "Toplam Ciro", "Toplam Kâr"]
            sayfa_ekle("Finansal Rapor", finansal, basliklar_finans)

            # ---------------------------------------------------------
            # 5. SEKME: EN ÇOK SATANLAR
            # ---------------------------------------------------------
            bestseller = self.veritabani.en_cok_satanlari_getir('tumu')
            basliklar_best = ["Kod", "Ürün Adı", "Kategori", "Toplam Satış Adedi", "Birim"]
            sayfa_ekle("En Çok Satanlar", bestseller, basliklar_best)

            # Dosyayı Kaydet
            wb.save(dosya_yolu)
            self.status_bar.showMessage(f"Tüm veriler başarıyla '{dosya_yolu}' dosyasına kaydedildi.", 5000)
            QMessageBox.information(self, "Başarılı",
                                    "Rapor oluşturuldu!\nDosya içeriğinde ayrı sekmeler halinde tüm verileri bulabilirsiniz.")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulurken hata meydana geldi:\n{str(e)}")


# =============================================================================
# 6. LOGIN AND SETUP WINDOWS
# =============================================================================

# --- BURADAKİ DEĞİŞİKLİK: QWidget yerine QDialog yapıldı ---
class BaseAuthWindow(QDialog):
    """
    Giriş, Kayıt ve Şifre Değiştirme ekranları için temel (parent) sınıf.
    Ortak tasarım özelliklerini (Logo, ortalama, tema) barındırır.
    """

    def __init__(self, title, size=(400, 300)):
        super().__init__()
        self.setWindowTitle(title)

        # 1. Boyutlandırma ve İsimlendirme
        self.setFixedSize(*size)
        self.setObjectName("authWindow")

        # Giriş ekranı her zaman KOYU tema olsun (Kullanıcı isteği)
        self.setStyleSheet(stil_olustur(True))

        # 2. Pencereyi Ekranın Ortasına Taşıma
        frame_geo = self.frameGeometry()
        screen = self.screen()
        if screen:
            center_point = screen.availableGeometry().center()
            frame_geo.moveCenter(center_point)
            self.move(frame_geo.topLeft())

        # 3. Ana Düzen (Layout)
        self.ana_layout = QVBoxLayout(self)
        # GÜNCELLEME 1: Ana pencerenin üst boşluğunu (ikinci parametre) azalttık (20 -> 5)
        self.ana_layout.setContentsMargins(20, 5, 20, 20)

        # 4. Logo Kısmı
        logo_layout = QVBoxLayout()
        # GÜNCELLEME 2: Logonun kendi üst boşluğunu sıfırladık (10 -> 0)
        logo_layout.setContentsMargins(0, 0, 0, 20)

        logo_icon = QLabel()
        logo_icon.setObjectName("logoIcon")
        try:
            logo_path = dosya_yolunu_bul("StockFlow_Logo.png")
            logo_pixmap = QPixmap(logo_path)
            if not logo_pixmap.isNull():
                logo_icon.setPixmap(logo_pixmap.scaled(400, 150, Qt.AspectRatioMode.KeepAspectRatio,
                                                       Qt.TransformationMode.SmoothTransformation))
                logo_icon.setAlignment(Qt.AlignmentFlag.AlignCenter)
            else:
                raise FileNotFoundError
        except FileNotFoundError:
            logo_icon.setText("STOCKFLOW")
            logo_icon.setStyleSheet(
                "font-size: 32px; font-weight: bold; color: #3b82f6; qproperty-alignment: AlignCenter;")

        logo_layout.addWidget(logo_icon)
        self.ana_layout.addLayout(logo_layout)


class IlkKurulumPenceresi(BaseAuthWindow):
    """
    Veritabanında hiç kullanıcı yoksa açılan ilk kurulum ve yönetici oluşturma ekranı.
    """
    kurulum_tamamlandi = pyqtSignal()

    def __init__(self, veritabani_yoneticisi):
        super().__init__("StockFlow - İlk Kurulum", (400, 350))
        self.veritabani = veritabani_yoneticisi
        duzen = QFormLayout()
        duzen.setContentsMargins(20, 0, 20, 0)
        duzen.setVerticalSpacing(15)
        self.ana_layout.addWidget(QLabel("<h3>Yönetici Hesabı Oluştur</h3>", alignment=Qt.AlignmentFlag.AlignCenter))
        self.k_adi = QLineEdit()
        self.k_adi.setPlaceholderText("Kullanıcı Adı")
        self.sifre = QLineEdit()
        self.sifre.setPlaceholderText("Şifre")
        self.sifre.setEchoMode(QLineEdit.EchoMode.Password)
        self.sifre_t = QLineEdit()
        self.sifre_t.setPlaceholderText("Şifre Tekrar")
        self.sifre_t.setEchoMode(QLineEdit.EchoMode.Password)
        duzen.addRow("Kullanıcı Adı:", self.k_adi)
        duzen.addRow("Şifre:", self.sifre)
        duzen.addRow("Şifre Tekrar:", self.sifre_t)
        self.ana_layout.addLayout(duzen)
        btn = QPushButton("Hesabı Oluştur")
        btn.clicked.connect(self.hesap_olustur)
        self.ana_layout.addWidget(btn)
        self.ana_layout.addStretch()
        self.sifre_t.returnPressed.connect(self.hesap_olustur)

    def hesap_olustur(self):
        k, s, st = self.k_adi.text().strip(), self.sifre.text(), self.sifre_t.text()
        if not k or not s: QMessageBox.warning(self, "Hata", "Alanlar boş bırakılamaz."); return
        if s != st: QMessageBox.warning(self, "Hata", "Şifreler eşleşmiyor."); return
        b, m = self.veritabani.kullanici_ekle(k, s)
        if b:
            QMessageBox.information(self, "Başarılı", "Yönetici hesabı oluşturuldu. Lütfen giriş yapın.")
            self.kurulum_tamamlandi.emit()
            self.close()
        else:
            QMessageBox.critical(self, "Hata", m)


class GirisPenceresi(BaseAuthWindow):
    """
    Standart kullanıcı giriş ekranı.
    Kullanıcı adı ve şifre doğrulaması yapar.
    """
    login_basarili = pyqtSignal(str)
    degistirme_penceresi_iste = pyqtSignal()
    yeni_kullanici_iste = pyqtSignal()

    def __init__(self, veritabani_yoneticisi):
        super().__init__("StockFlow - Giriş Yap", (400, 380))
        self.veritabani = veritabani_yoneticisi
        duzen = QVBoxLayout()
        duzen.setContentsMargins(20, 0, 20, 0)
        duzen.setSpacing(5)
        self.k_adi = QLineEdit()
        self.sifre = QLineEdit()
        self.sifre.setEchoMode(QLineEdit.EchoMode.Password)
        duzen.addWidget(QLabel("Kullanıcı Adı:"))
        duzen.addWidget(self.k_adi)
        duzen.addSpacing(10)
        duzen.addWidget(QLabel("Şifre:"))
        duzen.addWidget(self.sifre)
        self.ana_layout.addLayout(duzen)
        login_btn = QPushButton("Giriş Yap")
        login_btn.clicked.connect(self.login_kontrol)
        self.ana_layout.addWidget(login_btn)
        alt_buton_layout = QHBoxLayout()
        alt_buton_layout.setContentsMargins(20, 10, 20, 0)
        yeni_kullanici_btn = QPushButton("Hesap Oluştur")
        yeni_kullanici_btn.setObjectName("linkButton")
        yeni_kullanici_btn.clicked.connect(self.yeni_kullanici_iste.emit)
        degistir_btn = QPushButton("Bilgileri Değiştir")
        degistir_btn.setObjectName("linkButton")
        degistir_btn.clicked.connect(self.degistirme_penceresi_iste.emit)
        alt_buton_layout.addWidget(yeni_kullanici_btn)
        alt_buton_layout.addStretch()
        alt_buton_layout.addWidget(degistir_btn)
        self.ana_layout.addLayout(alt_buton_layout)
        self.ana_layout.addStretch()
        self.sifre.returnPressed.connect(self.login_kontrol)
        self.k_adi.returnPressed.connect(self.sifre.setFocus)

    def login_kontrol(self):
        k, s = self.k_adi.text(), self.sifre.text()
        if self.veritabani.kullanici_dogrula(k, s):
            self.login_basarili.emit(k)
            self.close()
        else:
            QMessageBox.warning(self, "Hata", "Kullanıcı adı veya şifre hatalı.")
            self.sifre.clear()
            self.sifre.setFocus()


class KullaniciDegistirPenceresi(BaseAuthWindow):
    """
    Mevcut kullanıcının şifresini veya adını değiştirmek için kullandığı form.
    Güvenlik için eski şifreyi de ister.
    """

    def __init__(self, veritabani_yoneticisi, mevcut_kullanici_adi=None):
        super().__init__("Bilgileri Değiştir", (450, 480))
        self.veritabani = veritabani_yoneticisi
        duzen = QFormLayout()
        duzen.setContentsMargins(20, 0, 20, 0)
        duzen.setVerticalSpacing(15)
        duzen.addRow(QLabel("<h3>Mevcut Bilgileri Doğrula</h3>"))
        self.e_kadi = QLineEdit()
        if mevcut_kullanici_adi:
            self.e_kadi.setText(mevcut_kullanici_adi)
            self.e_kadi.setEnabled(False)
        else:
            self.e_kadi.setPlaceholderText("Mevcut Kullanıcı Adı")
        self.e_sifre = QLineEdit()
        self.e_sifre.setPlaceholderText("Mevcut Şifre")
        self.e_sifre.setEchoMode(QLineEdit.EchoMode.Password)
        duzen.addRow("Mevcut Kullanıcı:", self.e_kadi)
        duzen.addRow("Mevcut Şifre:", self.e_sifre)
        duzen.addRow(QLabel("<h3>Yeni Bilgileri Girin</h3>"))
        self.y_kadi = QLineEdit()
        self.y_kadi.setPlaceholderText("Yeni Kullanıcı Adı")
        self.y_sifre = QLineEdit()
        self.y_sifre.setPlaceholderText("Yeni Şifre")
        self.y_sifre.setEchoMode(QLineEdit.EchoMode.Password)
        self.y_sifre_t = QLineEdit()
        self.y_sifre_t.setPlaceholderText("Yeni Şifre Tekrar")
        self.y_sifre_t.setEchoMode(QLineEdit.EchoMode.Password)
        duzen.addRow("Yeni Kullanıcı Adı:", self.y_kadi)
        duzen.addRow("Yeni Şifre:", self.y_sifre)
        duzen.addRow("Yeni Şifre Tekrar:", self.y_sifre_t)
        self.ana_layout.addLayout(duzen)

        self.buttonBox = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)

        # Butonların üzerindeki yazıları değiştiriyoruz
        self.buttonBox.button(QDialogButtonBox.StandardButton.Ok).setText("Değişiklikleri Onayla")
        self.buttonBox.button(QDialogButtonBox.StandardButton.Cancel).setText("Geri Dön")

        # Tıklama olaylarını bağlıyoruz
        self.buttonBox.accepted.connect(self.bilgileri_degistir)  # Onayla'ya basınca kaydet
        self.buttonBox.rejected.connect(self.reject)  # Geri'ye basınca pencereyi kapat

        # Ekrana ekliyoruz
        self.ana_layout.addWidget(self.buttonBox)
        self.ana_layout.addStretch()

        # Enter tuşuna basınca onaylama çalışsın
        self.y_sifre_t.returnPressed.connect(self.bilgileri_degistir)

    def bilgileri_degistir(self):
        ek, es = self.e_kadi.text().strip(), self.e_sifre.text()
        yk, ys, yst = self.y_kadi.text().strip(), self.y_sifre.text(), self.y_sifre_t.text()
        if not all([ek, es, yk, ys]): QMessageBox.warning(self, "Hata", "Tüm alanlar doldurulmalıdır."); return
        if ys != yst: QMessageBox.warning(self, "Hata", "Yeni şifreler eşleşmiyor."); return
        if not self.veritabani.kullanici_dogrula(ek, es): QMessageBox.warning(self, "Hata",
                                                                              "Mevcut bilgiler yanlış."); return
        b, m = self.veritabani.kullanici_bilgilerini_guncelle(ek, yk, ys)
        if b:
            QMessageBox.information(self, "Başarılı",
                                    m + "\nDeğişikliklerin geçerli olması için lütfen yeniden giriş yapın.");
            self.close()
        else:
            QMessageBox.warning(self, "Hata", m)


# =============================================================================
# 7. APPLICATION CONTROLLER
# =============================================================================

class AnaKontrolcu:
    """
    Uygulamanın akışını (Flow Controller) yöneten ana sınıf.
    Hangi pencerenin (Giriş mi, Ana Ekran mı, Kurulum mu) gösterileceğine karar verir.
    """

    def __init__(self):
        self.veritabani = VeritabaniYoneticisi()
        self.login_penceresi = self.ana_pencere = self.mevcut_pencere = None
        self.degistirme_penceresi = None
        self.yeni_kullanici_penceresi = None

    def baslat(self):
        if self.veritabani.kullanici_sayisi_getir() == 0:
            self.mevcut_pencere = IlkKurulumPenceresi(self.veritabani)
            self.mevcut_pencere.kurulum_tamamlandi.connect(self.login_penceresini_goster)
        else:
            self.login_penceresini_goster()
        self.mevcut_pencere.show()

    def login_penceresini_goster(self, _=None):
        if self.mevcut_pencere and isinstance(self.mevcut_pencere, IlkKurulumPenceresi):
            self.mevcut_pencere.close()
        self.login_penceresi = GirisPenceresi(self.veritabani)
        self.login_penceresi.login_basarili.connect(self.ana_pencereyi_goster)
        self.login_penceresi.degistirme_penceresi_iste.connect(self.degistirme_penceresini_goster)
        self.login_penceresi.yeni_kullanici_iste.connect(self.yeni_kullanici_penceresi_goster)
        self.mevcut_pencere = self.login_penceresi
        self.mevcut_pencere.show()

    def ana_pencereyi_goster(self, kullanici_adi):
        try:
            self.ana_pencere = AnaPencere(kullanici_adi, self.veritabani)
            self.ana_pencere.cikis_istendi.connect(self.oturum_kapat_ve_giris_yap)
            self.ana_pencere.showMaximized()
            self.bildirim_gonder_kontrolu()
            if self.mevcut_pencere: self.mevcut_pencere.close()
            if self.degistirme_penceresi:
                self.degistirme_penceresi.close()
                self.degistirme_penceresi = None
            if self.yeni_kullanici_penceresi:
                self.yeni_kullanici_penceresi.close()
                self.yeni_kullanici_penceresi = None
            self.mevcut_pencere = self.login_penceresi = None
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(None, "Kritik Hata", f"Ana pencere açılırken bir hata oluştu:\n{str(e)}")
            print(f"HATA: {e}")
            pass

    def oturum_kapat_ve_giris_yap(self):
        """Ana pencere kapandığında giriş ekranını tekrar açar."""
        # Ana pencere zaten kapanıyor ama referansı temizleyelim
        self.ana_pencere = None
        # Giriş ekranını tekrar çağır
        self.login_penceresini_goster()

    def degistirme_penceresini_goster(self):
        if self.login_penceresi and self.login_penceresi.isVisible(): self.login_penceresi.hide()
        if self.degistirme_penceresi and self.degistirme_penceresi.isVisible():
            self.degistirme_penceresi.activateWindow()
        else:
            self.degistirme_penceresi = KullaniciDegistirPenceresi(self.veritabani)
            self.degistirme_penceresi.exec()
        if self.login_penceresi: self.login_penceresi.show()

    def yeni_kullanici_penceresi_goster(self):
        if self.login_penceresi and self.login_penceresi.isVisible(): self.login_penceresi.hide()
        if self.yeni_kullanici_penceresi and self.yeni_kullanici_penceresi.isVisible():
            self.yeni_kullanici_penceresi.activateWindow()
        else:
            self.yeni_kullanici_penceresi = YeniKullaniciDialog(self.veritabani)
            self.yeni_kullanici_penceresi.exec()
        if self.login_penceresi: self.login_penceresi.show()

    def bildirim_gonder_kontrolu(self):
        if not PLYER_MUMKUN: return
        try:
            dusuk_stok_adedi = len(self.veritabani.dusuk_stok_urunleri_getir())
            if dusuk_stok_adedi > 0:
                notification.notify(title='StockFlow - Düşük Stok Uyarısı',
                                    message=f'Stok seviyesi kritik olan {dusuk_stok_adedi} adet ürününüz var.',
                                    app_name='StockFlow', timeout=10)
        except Exception as e:
            print(f"Bildirim hatası: {e}")


# =============================================================================
# 8. STYLESHEET AND EXECUTION
# =============================================================================

def stil_olustur(koyu_tema=True):
    if koyu_tema:
        # --- KOYU TEMA RENKLERİ ---
        c = {
            "bg_main": "#20293a",  # Ana Arka Plan
            "bg_sidebar": "#0f172a",  # Yan Menü Arka Planı
            "bg_card": "#334155",  # Kart/Kutu Arka Planı
            "bg_header": "#20293a",  # Üst Bar
            "bg_hover": "#3e4c63",  # Mouse üzerine gelince
            "bg_input": "#334155",  # Input alanları
            "text_main": "#e2e8f0",  # Ana Metin
            "text_muted": "#94a3b8",  # Yan Metin
            "border": "#475569",  # Kenarlıklar
            "accent": "#3b82f6",  # Vurgu Rengi (Mavi)
            "success": "#10b981",  # Yeşil
            "danger": "#ef4444",  # Kırmızı
            "table_header": "#1e293b",  # Tablo Başlığı
            "scroll_bg": "#20293a",  # Scrollbar arkası
            "scroll_handle": "#475569",  # Scrollbar tutacağı
            "text_sidebar": "#94a3b8"  # Sidebar metin rengi
        }
    else:
        # --- AÇIK TEMA RENKLERİ ---
        c = {
            "bg_main": "#f1f5f9",  # Çok açık gri (neredeyse beyaz)
            "bg_sidebar": "#0f172a",  # KOYU SIDEBAR (Logo görünürlüğü için)
            "bg_card": "#ffffff",  # Beyaz Kartlar
            "bg_header": "#ffffff",  # Beyaz Header
            "bg_hover": "#e2e8f0",  # Açık gri hover
            "bg_input": "#ffffff",  # Beyaz input
            "text_main": "#1e293b",  # Koyu lacivert/siyah metin
            "text_muted": "#64748b",  # Gri metin
            "border": "#cbd5e1",  # Açık gri kenarlık
            "accent": "#3b82f6",  # Mavi (Aynı kalabilir)
            "success": "#10b981",  # Yeşil
            "danger": "#ef4444",  # Kırmızı
            "table_header": "#f8fafc",  # Çok açık gri tablo başlığı
            "scroll_bg": "#f1f5f9",
            "scroll_handle": "#cbd5e1",
            "text_sidebar": "#cbd5e1"  # Sidebar metni (Koyu zemin üstüne açık gri)
        }

    return f"""
    /* Global Reset & Fonts */
    * {{ font-family: 'Segoe UI', 'Roboto', sans-serif; color: {c['text_main']}; selection-background-color: {c['accent']}; selection-color: #ffffff; }}

    /* Main Backgrounds */
    QMainWindow, QWidget#central_widget {{ background-color: {c['bg_main']}; }}
    QDialog, QWidget#authWindow {{ background-color: {c['bg_main']}; border: 1px solid {c['border']}; border-radius: 8px; }}
    QMessageBox {{ background-color: {c['bg_main']}; color: {c['text_main']}; }}

    /* Sidebar */
    QFrame#sidebarFrame {{ background-color: {c['bg_sidebar']}; border-right: 1px solid {c['border']}; }}
    QFrame#sidebarFrame QLabel {{ color: {c['text_sidebar']}; font-size: 11px; font-weight: 700; text-transform: uppercase; padding-left: 12px; padding-top: 20px; letter-spacing: 0.5px; }}
    QLabel#logoIcon {{ qproperty-alignment: 'AlignCenter'; min-height: 50px; padding: 10px; margin: 0; }}
    QLabel#logoText {{ font-size: 22px; font-weight: 800; color: {c['text_main']}; padding: 0; letter-spacing: 0.5px; }}

    /* Navigation List */
    QListWidget#sidebarNav {{ border: none; background: transparent; outline: none; margin-top: 10px; }}
    QListWidget#sidebarNav::item {{ padding: 12px 16px; border-radius: 8px; color: {c['text_sidebar']}; font-weight: 500; margin: 4px 12px; transition: all 0.2s; }}
    QListWidget#sidebarNav::item:hover {{ background-color: {c['bg_hover']}; color: {c['text_main']}; }}
    QListWidget#sidebarNav::item:selected {{ background-color: {c['accent']}; color: #ffffff; font-weight: 600; border-left: 4px solid #60a5fa; }}

    /* Buttons */
    QPushButton {{ background-color: {c['accent']}; color: white; border: none; border-radius: 6px; padding: 8px 16px; font-weight: 600; font-size: 13px; }}
    QPushButton:hover {{ background-color: #2563eb; }}
    QPushButton:pressed {{ background-color: #1d4ed8; }}
    QPushButton:disabled {{ background-color: {c['border']}; color: {c['text_muted']}; }}

    QPushButton#sourceButton {{ background: {c['bg_sidebar']}; border: 1px solid {c['border']}; color: {c['text_sidebar']}; padding: 12px; text-align: left; border-radius: 8px; font-size: 12px; font-weight: 500; margin: 10px; }}
    QPushButton#headerButton {{ background: transparent; border: 1px solid {c['border']}; color: {c['text_muted']}; font-weight: 500; padding: 6px 12px; }}
    QPushButton#headerButton:hover {{ background: {c['bg_card']}; color: {c['text_main']}; border-color: {c['border']}; }}

    QPushButton#filterBtn {{ background: {c['bg_input']}; border: 1px solid {c['border']}; color: {c['text_main']}; }}
    QPushButton#filterBtn:hover {{ background: {c['bg_hover']}; border-color: {c['text_muted']}; }}
    QPushButton#filterBtn[filtered="true"] {{ background: {c['accent']}; color: white; border: none; }}

    QPushButton#yeniUrunBtn {{ background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 {c['success']}, stop:1 #059669); }}
    QPushButton#yeniUrunBtn:hover {{ background: #059669; }}

    QPushButton#menuButton {{ background: transparent; border: 1px solid transparent; border-radius: 4px; font-weight: 900; font-size: 18px; color: {c['text_muted']}; padding: 0px 8px; margin: 0; line-height: 10px; }}
    QPushButton#menuButton:hover {{ background: {c['bg_hover']}; color: {c['text_main']}; border-color: {c['border']}; }}

    QPushButton#linkButton {{ background: transparent; color: {c['accent']}; text-decoration: none; font-weight: 500; padding: 4px; }}
    QPushButton#linkButton:hover {{ color: #93c5fd; text-decoration: underline; }}

    QPushButton#closeBtn {{ background-color: transparent; color: {c['text_muted']}; border: 1px solid {c['border']}; }}
    QPushButton#closeBtn:hover {{ background-color: {c['danger']}; color: white; }}

    /* Header & Content */
    QFrame#contentContainer {{ background: {c['bg_main']}; }}
    QFrame#headerBar {{ background-color: {c['bg_header']}; border-bottom: 1px solid {c['border']}; }}
    QLabel#pageTitle {{ font-size: 24px; font-weight: 700; color: {c['text_main']}; padding: 0; }}
    QLabel#pageSubtitle {{ font-size: 14px; color: {c['text_muted']}; }}

    /* Cards */
    QFrame#metricCard {{
        background-color: {c['bg_card']};
        border: 1px solid {c['border']};
        border-radius: 12px;
        padding: 20px;
    }}
    QFrame#metricCard[clickable="true"]:hover {{
        background-color: {c['bg_hover']};
        border-color: {c['accent']};
        cursor: pointer;
    }}

    /* Chart Cards */
    QFrame#chartCard {{ background-color: {c['bg_card']}; border-radius: 12px; border: 1px solid {c['border']}; }}
    QFrame#chartCard:hover {{ border-color: {c['accent']}; background-color: {c['bg_hover']}; cursor: pointer; }}

    QLabel#metricTitle {{ font-size: 12px; font-weight: 600; color: {c['text_muted']}; text-transform: uppercase; letter-spacing: 0.5px; }}
    QLabel#metricValue {{ font-size: 24px; font-weight: 800; color: {c['text_main']}; padding: 2px 0; }}
    QLabel#metricValue[lowStock="true"] {{ color: {c['danger']}; }}
    QLabel#metricUnit {{ font-size: 11px; color: {c['text_muted']}; margin-left: 4px; padding-bottom: 4px; }}

    /* Tour Info Box */
    QLabel#tourTitle {{ color: {c['text_main']}; font-weight: bold; font-size: 16px; }}
    QLabel#tourDesc {{ color: {c['text_main']}; font-size: 13px; }}

    /* Search & Inputs */
    QFrame#searchContainer {{ background-color: {c['bg_input']}; border: 1px solid {c['border']}; border-radius: 8px; }}
    QLineEdit#searchInput {{ border: none; background: transparent; padding: 10px; font-size: 14px; color: {c['text_main']}; }}

    QLineEdit, QComboBox, QDateEdit {{ background-color: {c['bg_input']}; border: 1px solid {c['border']}; border-radius: 6px; padding: 8px; color: {c['text_main']}; font-size: 13px; }}
    QLineEdit:focus, QComboBox:focus, QDateEdit:focus {{ border-color: {c['accent']}; background-color: {c['bg_hover']}; }}
    QLineEdit::placeholder {{ color: {c['text_muted']}; }}

    /* Combo Box Dropdown (Popup) Fix */
    QComboBox QAbstractItemView {{
        background-color: {c['bg_card']};
        color: {c['text_main']};
        selection-background-color: {c['accent']};
        selection-color: #ffffff;
        border: 1px solid {c['border']};
        outline: none;
    }}

    /* Date Edit Calendar Fix */
    QCalendarWidget QWidget {{ background-color: {c['bg_card']}; color: {c['text_main']}; }}
    QCalendarWidget QToolButton {{ color: {c['text_main']}; icon-size: 24px; }}
    QCalendarWidget QMenu {{ background-color: {c['bg_card']}; color: {c['text_main']}; }}
    QCalendarWidget QSpinBox {{ background-color: {c['bg_input']}; color: {c['text_main']}; }}
    QCalendarWidget QAbstractItemView:enabled {{ 
        color: {c['text_main']}; 
        background-color: {c['bg_card']}; 
        selection-background-color: {c['accent']}; 
        selection-color: white; 
    }}

    /* Tables */
    QTableWidget {{ background-color: {c['bg_card']}; border: 1px solid {c['border']}; border-radius: 8px; gridline-color: {c['border']}; outline: none; }}
    QTableWidget::item {{ padding: 12px 8px; border-bottom: 1px solid {c['border']}; color: {c['text_main']}; }}
    QTableWidget::item:selected {{ background-color: {c['accent']}; color: #ffffff; }}
    QHeaderView::section {{ background-color: {c['table_header']}; padding: 8px 6px; border: none; border-bottom: 2px solid {c['border']}; font-weight: 700; color: {c['text_muted']}; text-transform: uppercase; font-size: 11px; }}
    QTableWidget QTableCornerButton::section {{ background-color: {c['table_header']}; border: none; }}

    /* Menus */
    QMenu {{ background: {c['bg_card']}; border: 1px solid {c['border']}; padding: 4px; border-radius: 6px; }}
    QMenu::item {{ padding: 6px 24px 6px 12px; border-radius: 4px; color: {c['text_main']}; }}
    QMenu::item:selected {{ background: {c['accent']}; color: white; }}
    QMenu::separator {{ height: 1px; background: {c['border']}; margin: 4px 0; }}

    /* ScrollArea & Dashboard Content */
    QScrollArea {{ background-color: transparent; border: none; }}
    QWidget#dashboardContent {{ background-color: {c['bg_main']}; }}

    /* Scrollbars */
    QScrollBar:vertical {{ border: none; background: {c['scroll_bg']}; width: 10px; margin: 0; }}
    QScrollBar::handle:vertical {{ background: {c['scroll_handle']}; min-height: 20px; border-radius: 5px; }}
    QScrollBar::handle:vertical:hover {{ background: {c['text_muted']}; }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0px; }}

    QScrollBar:horizontal {{ border: none; background: {c['scroll_bg']}; height: 10px; margin: 0; }}
    QScrollBar::handle:horizontal {{ background: {c['scroll_handle']}; min-width: 20px; border-radius: 5px; }}

    /* Status Bar */
    QStatusBar {{ background: {c['bg_sidebar']}; border-top: 1px solid {c['border']}; color: {c['text_muted']}; padding: 4px; }}
    """


if __name__ == "__main__":
    import sys
    import os
    from PyQt6.QtWidgets import QApplication
    from PyQt6.QtGui import QIcon


    # --- GLOBAL HATA YAKALAMA ---
    def exception_hook(exctype, value, traceback):
        print(f"Hata: {value}")
        sys.__excepthook__(exctype, value, traceback)


    sys.excepthook = exception_hook

    # 1. Uygulamayı Başlat
    # --- DPI ÖLÇEKLENDİRME AYARLARI (Scaling Fix) ---
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    QApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.PassThrough)

    app = QApplication(sys.argv)
    app.setStyleSheet(stil_olustur(False))

    # 2. İşletim Sistemi Kontrolü ve İkon Yolu Belirleme
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # İkon dosyasını belirle (Önce .ico, yoksa .png dene)
    icon_path = os.path.join(script_dir, "logo.png")

    # 3. WINDOWS'A ÖZEL AYAR (Sadece Windows ise çalışır)
    # os.name 'nt' ise Windows demektir.
    if os.name == 'nt':
        import ctypes

        myappid = 'stockflow.stok.yonetimi.v1.0'  # Benzersiz kimlik
        try:
            # Bu satır Windows'ta görev çubuğu ikonunu düzeltir
            ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
        except Exception:
            pass

    # 4. İKONU UYGULAMA (Tüm Sistemler İçin Ortak)
    if os.path.exists(icon_path):
        app_icon = QIcon(icon_path)
        app.setWindowIcon(app_icon)



    # 5. Ana Pencereyi Başlat
    kontrolcu = AnaKontrolcu()
    kontrolcu.baslat()

    sys.exit(app.exec())
